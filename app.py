# ==============================================
# 📄 CONFIGURACIONES INICIALES
# ==============================================

# Importar configuración de sesión corregida
import time
import certifi
# --- Imports estándar ---
import os
import sys
import logging
import tempfile
import zipfile
import secrets
import traceback
from datetime import datetime, time, timedelta

# --- Imports de terceros ---
import certifi
import boto3
import openpyxl
from bson import ObjectId
from dotenv import load_dotenv
from pymongo import MongoClient
from pymongo.server_api import ServerApi
from flask import Flask, render_template, request, redirect, send_from_directory, url_for, flash, session, jsonify, send_file, abort, current_app
from flask_pymongo import PyMongo
from flask_mail import Mail, Message
from flask_session import Session
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
from redis import Redis
from app.decorators import admin_required, login_required

# Importar otros módulos de utilidad según sea necesario
import hashlib
from openpyxl import Workbook  # <--- Importación correcta
from app.extensions import init_extensions

print('DEBUG IMPORT: login_required y admin_required importados desde app.decorators')

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in {"png", "jpg", "jpeg", "gif"}

def eliminar_archivo_imagen(ruta):
    if not ruta:
        return
    if ruta.startswith("s3://"):
        s3_parts = ruta[5:].split('/', 1)
        if len(s3_parts) == 2:
            bucket_name, object_key = s3_parts
            if bucket_name == current_app.S3_BUCKET_NAME:
                try:
                    current_app.s3_client.delete_object(Bucket=bucket_name, Key=object_key)
                except Exception as e:
                    current_app.logger.error(f"Error eliminando imagen de S3: {e}")
    else:
        local_path = os.path.join(current_app.config["UPLOAD_FOLDER"], os.path.basename(ruta))
        if os.path.exists(local_path):
            try:
                os.remove(local_path)
            except Exception as e:
                current_app.logger.error(f"Error eliminando imagen local: {e}")

def get_current_spreadsheet():
    selected_table = session.get("selected_table")
    if not selected_table:
        return None
    return os.path.join(current_app.config["UPLOAD_FOLDER"], selected_table)

def leer_datos_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    hoja = wb.active
    headers = [cell.value for cell in next(hoja.iter_rows(min_row=1, max_row=1))]
    data = []
    for row in hoja.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))
    wb.close()
    return data

# ==============================================
# 📄 CONFIGURACIÓN INICIAL DE FLASK
# ==============================================

# --- Cargar variables de entorno ---
load_dotenv()

def create_app():
    print("DEBUG: create_app ejecutado")
    """Crea y configura la aplicación Flask"""
    ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
    TEMPLATE_DIR = os.path.join(ROOT_DIR, "app", "templates")
    app = Flask(__name__, template_folder=TEMPLATE_DIR)

    # --- Registrar blueprint de imágenes ---
    from app.routes.images_routes import images_bp
    app.register_blueprint(images_bp)
    print("DEBUG: images_bp registrado en app")

    # Cargar configuración desde config.py primero
    app.config.from_object('config.Config')
    
    # CONFIGURACIÓN DE SESIÓN DIRECTA - Garantiza funcionamiento correcto
    # Esta configuración tiene prioridad sobre cualquier otra
    app.config['SESSION_TYPE'] = 'filesystem'
    app.config['SESSION_FILE_DIR'] = os.path.join(ROOT_DIR, 'flask_session')
    app.config['SESSION_COOKIE_NAME'] = 'edefrutos2025_session'
    app.config['SESSION_COOKIE_SECURE'] = False  # IMPORTANTE: Deshabilitado para desarrollo (permite HTTP)
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
    app.config['SESSION_PERMANENT'] = True
    app.config['PERMANENT_SESSION_LIFETIME'] = 86400  # 24 horas en segundos
    app.config['SESSION_REFRESH_EACH_REQUEST'] = True
    app.config['SESSION_USE_SIGNER'] = False  # Deshabilitado para simplificar depuración
    
    # Usar clave secreta fija para desarrollo (garantiza persistencia entre reinicios)
    if not os.environ.get('SECRET_KEY'):
        app.config['SECRET_KEY'] = 'desarrollo_clave_secreta_fija_12345'
    else:
        app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY')
    
    # Establecer también como atributo directo para mayor compatibilidad
    app.secret_key = app.config['SECRET_KEY']
    
    app.logger.info("✅ Configuración de sesión aplicada directamente en app.py")
    
    # Asegurarse de que el valor de SESSION_FILE_DIR sea absoluto
    if 'SESSION_FILE_DIR' in app.config and not os.path.isabs(app.config['SESSION_FILE_DIR']):
        app.config['SESSION_FILE_DIR'] = os.path.join(ROOT_DIR, app.config['SESSION_FILE_DIR'])
    
    # Asegurarse de que el directorio de sesiones existe
    os.makedirs(app.config['SESSION_FILE_DIR'], exist_ok=True)
    
    # Inicializar Flask-Session (debe estar después de configurar app.config)
    Session(app)
    app.logger.info("✅ Flask-Session inicializado correctamente")
    
    # Configurar sesión permanente por defecto
    @app.before_request
    def make_session_permanent():
        session.permanent = True
    
    # Inicializar la conexión global a MongoDB (para funciones legacy y modelos)
    try:
        from app.database import initialize_db, get_mongo_client, get_mongo_db
        initialize_db(app)
        app.logger.info("✅ Conexión global a MongoDB inicializada (initialize_db)")
        # Refuerza la asignación de app.db y colecciones SIEMPRE tras inicializar
        client = get_mongo_client()
        db = get_mongo_db()
        app.mongo_client = client
        app.db = db
        if db is not None:
            app.users_collection = db["users"]
            app.resets_collection = db["password_resets"]
            app.catalog_collection = db["67b8c24a7fdc72dd4d8703cf"]
            app.spreadsheets_collection = db["spreadsheets"]
        else:
            app.users_collection = None
            app.resets_collection = None
            app.catalog_collection = None
            app.spreadsheets_collection = None
    except Exception as e:
        app.logger.error(f"❌ Error inicializando la conexión global a MongoDB: {e}")
        app.db = None
        app.users_collection = None
        app.resets_collection = None
        app.catalog_collection = None
        app.spreadsheets_collection = None

    # =================== FUNCIONES AUXILIARES ===================
    # (Eliminar las definiciones internas de eliminar_archivo_imagen, get_current_spreadsheet y leer_datos_excel aquí)

    # =================== RUTAS Y HANDLERS ===================
    # (Las rutas pueden ahora usar current_app.users_collection, etc.)
    # Ejemplo de uso en una ruta:
    # usuario = current_app.users_collection.find_one({...})
    # s3 = current_app.s3_client
    # bucket = current_app.S3_BUCKET_NAME
    # ...

    # --- Versión CSS automática para evitar caché ---
    CSS_VERSION = datetime.now().strftime("%Y%m%d%H%M%S")

    # --- Función para saber si un valor es flotante ---
    def is_float(value):
        try:
            float(value)
            return True
        except (ValueError, TypeError):
            return False

    # --- Inyectar variables globales en todas las plantillas ---
    @app.context_processor
    def inject_template_vars():
        return {
            'now': datetime.now(),
            'css_version': CSS_VERSION
        }

    # --- Configurar archivos estáticos personalizados ---
    @app.route('/static/<path:filename>')
    def custom_static(filename):
        try:
            return send_from_directory('static', filename)
        except Exception as e:
            app.logger.error(f"Error sirviendo archivo estático {filename}: {str(e)}")
            return str(e), 500

    # --- Decorador para rutas protegidas ---
    # def login_required(f):
    #     def wrapper(*args, **kwargs):
    #         if 'user_id' not in session:
    #             flash('Debes iniciar sesión primero.', 'warning')
    #             return redirect(url_for('auth.login'))
    #         return f(*args, **kwargs)
    #     wrapper.__name__ = f.__name__
    #     return wrapper

    # Registrar blueprints principales
    from app.routes.auth_routes import auth_bp
    from app.routes.main_routes import main_bp
    from app.routes.catalogs_routes import catalogs_bp
    from app.routes.catalog_images_routes import image_bp
    from app.routes.usuarios_routes import usuarios_bp
    from app.routes.admin_routes import admin_bp
    from app.routes.error_routes import errors_bp
    from app.routes.emergency_access import emergency_bp
    from app.routes.debug_routes import debug_bp  # Blueprint para diagnóstico de sesiones
    from app.routes.admin_diagnostic import admin_diagnostic_bp  # Blueprint para diagnóstico de administrador
    from app.routes.diagnostico import diagnostico_bp  # Blueprint para diagnóstico simplificado
    print("ANTES DE BLUEPRINTS", app.db)
    app.register_blueprint(auth_bp, url_prefix='')
    app.register_blueprint(main_bp)
    app.register_blueprint(catalogs_bp)
    app.register_blueprint(image_bp)
    app.register_blueprint(usuarios_bp)
    app.register_blueprint(admin_bp, url_prefix='/admin')
    app.register_blueprint(errors_bp)
    app.register_blueprint(emergency_bp)  # <-- REGISTRO DE EMERGENCIA
    print("DESPUÉS DE BLUEPRINTS", app.db)
    
    # Importar explícitamente los blueprints de diagnóstico
    try:
        from app.routes.debug_routes import debug_bp
        from app.routes.admin_diagnostic import admin_diagnostic_bp
        from app.routes.diagnostico import diagnostico_bp
        
        # Registrar blueprints de diagnóstico con sus prefijos URL
        app.register_blueprint(debug_bp)
        app.register_blueprint(admin_diagnostic_bp)
        app.register_blueprint(diagnostico_bp)
        
        app.logger.info("✅ Todos los blueprints de diagnóstico registrados correctamente")
    except Exception as e:
        print(f"Error durante el registro de blueprints: {e}")
        app.logger.error(f"❌ Error al registrar blueprints de diagnóstico: {e}")
        app.logger.error(traceback.format_exc())
    

    # Añadir rutas de prueba de sesión directamente en la aplicación principal
    @app.route('/prueba_sesion')
    def prueba_sesion():
        """Ruta simple para probar la persistencia de sesiones."""
        # Añadir un valor a la sesión
        session['prueba_timestamp'] = datetime.now().isoformat()
        session['contador'] = session.get('contador', 0) + 1
        session.modified = True
        
        # Preparar respuesta HTML
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Prueba de Sesión</title>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
        </head>
        <body>
            <div class="container mt-4">
                <h1>Prueba de Sesión</h1>
                
                <div class="alert alert-info">
                    <p>Esta página permite verificar que las sesiones están funcionando correctamente.</p>
                    <p>Si el contador aumenta al recargar la página, las sesiones están funcionando correctamente.</p>
                </div>
                
                <div class="card mb-4">
                    <div class="card-header bg-primary text-white">
                        <h5 class="mb-0">Estado de la Sesión</h5>
                    </div>
                    <div class="card-body">
                        <p><strong>Timestamp:</strong> {session.get('prueba_timestamp', 'No disponible')}</p>
                        <p><strong>Contador:</strong> {session.get('contador', 0)}</p>
                        <p><strong>Sesión Permanente:</strong> {session.permanent}</p>
                        <p><strong>ID de Sesión:</strong> {request.cookies.get(app.config.get('SESSION_COOKIE_NAME', ''), 'No disponible')}</p>
                        
                        <h6 class="mt-3">Contenido completo de la sesión:</h6>
                        <pre>{dict(session)}</pre>
                        
                        <div class="mt-3">
                            <a href="/prueba_sesion" class="btn btn-primary">Actualizar</a>
                            <a href="/prueba_sesion/limpiar" class="btn btn-danger">Limpiar sesión</a>
                            <a href="/prueba_sesion/test-cookie" class="btn btn-warning">Probar Cookie</a>
                        </div>
                    </div>
                </div>
                
                <div class="card mb-4">
                    <div class="card-header bg-info text-white">
                        <h5 class="mb-0">Configuración de Sesión</h5>
                    </div>
                    <div class="card-body">
                        <p><strong>SESSION_TYPE:</strong> {app.config.get('SESSION_TYPE', 'No configurado')}</p>
                        <p><strong>SESSION_FILE_DIR:</strong> {app.config.get('SESSION_FILE_DIR', 'No configurado')}</p>
                        <p><strong>SESSION_COOKIE_NAME:</strong> {app.config.get('SESSION_COOKIE_NAME', 'No configurado')}</p>
                        <p><strong>SESSION_COOKIE_SECURE:</strong> {app.config.get('SESSION_COOKIE_SECURE', False)}</p>
                        <p><strong>SESSION_COOKIE_HTTPONLY:</strong> {app.config.get('SESSION_COOKIE_HTTPONLY', True)}</p>
                        <p><strong>SESSION_COOKIE_SAMESITE:</strong> {app.config.get('SESSION_COOKIE_SAMESITE', 'No configurado')}</p>
                        <p><strong>SESSION_REFRESH_EACH_REQUEST:</strong> {app.config.get('SESSION_REFRESH_EACH_REQUEST', False)}</p>
                        <p><strong>SESSION_USE_SIGNER:</strong> {app.config.get('SESSION_USE_SIGNER', False)}</p>
                        <p><strong>PERMANENT_SESSION_LIFETIME:</strong> {app.config.get('PERMANENT_SESSION_LIFETIME', 'No configurado')}</p>
                        <p><strong>SECRET_KEY:</strong> {'*' * len(str(app.config.get('SECRET_KEY', '')))}</p>
                    </div>
                </div>
                
                <div class="card mb-4">
                    <div class="card-header bg-secondary text-white">
                        <h5 class="mb-0">Cookies Recibidas</h5>
                    </div>
                    <div class="card-body">
                        <pre>{request.cookies}</pre>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        return html
    
    @app.route('/prueba_sesion/limpiar')
    def prueba_sesion_limpiar():
        """Limpia la sesión actual."""
        session.clear()
        return redirect('/prueba_sesion')
        
    @app.route('/prueba_sesion/test-cookie')
    def prueba_sesion_test_cookie():
        """Prueba la configuración de cookies."""
        resp = redirect('/prueba_sesion')
        resp.set_cookie('test_cookie', 'valor_de_prueba', max_age=3600)
        return resp
    

    # Ruta de acceso directo a catálogos
    @app.route('/acceso_directo_catalogs')
    def acceso_directo_catalogs():
        # Establecer datos de sesión para el administrador
        session['logged_in'] = True
        session['email'] = 'admin@example.com'
        session['username'] = 'admin'
        session['role'] = 'admin'
        session['name'] = 'Administrador'
        
        app.logger.info("Sesión establecida para administrador mediante acceso directo")
        app.logger.info(f"Datos de sesión: {dict(session)}")
        
        # Redirigir a los catálogos
        return redirect(url_for('catalogs.list'))
    
    # Ruta de acceso directo para usuario normal
    @app.route('/acceso_directo_usuario')
    def acceso_directo_usuario():
        # Establecer datos de sesión para el usuario normal
        session['logged_in'] = True
        session['email'] = 'usuario@example.com'
        session['username'] = 'usuario'
        session['role'] = 'user'
        session['name'] = 'Usuario Normal'
        
        app.logger.info("Sesión establecida para usuario normal mediante acceso directo")
        app.logger.info(f"Datos de sesión: {dict(session)}")
        
        # Redirigir a los catálogos
        return redirect(url_for('catalogs.list'))

    def ensure_db():
        from app.database import get_mongo_db, get_mongo_client
        client = get_mongo_client()
        db = get_mongo_db()
        current_app.mongo_client = client
        current_app.db = db
        if db is not None:
            current_app.users_collection = db["users"]
            current_app.resets_collection = db["password_resets"]
            current_app.spreadsheets_collection = db["spreadsheets"]
        else:
            current_app.users_collection = None
            current_app.resets_collection = None
            current_app.spreadsheets_collection = None
    app.before_request(ensure_db)

    return app

app = create_app()

if __name__ == '__main__':
    # Leer el modo debug de la variable de entorno (por defecto False)
    debug_mode = os.getenv('FLASK_DEBUG', '0') in ('1', 'true', 'True')
    app.run(debug=debug_mode, host='0.0.0.0', port=5001)

# ==============================================
# 📄 INICIALIZACIÓN DE VARIABLES GLOBALES
# ==============================================

# Variables para carpetas
ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
SPREADSHEET_FOLDER = os.path.join(ROOT_DIR, "spreadsheets")
UPLOAD_FOLDER = os.path.join(ROOT_DIR, "imagenes_subidas")

# Extensiones de imagen permitidas
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif"}

# Versión de CSS para cache busting
CSS_VERSION = datetime.now().strftime("%Y%m%d%H%M%S")

# Asegurar que las carpetas necesarias existan
os.makedirs(SPREADSHEET_FOLDER, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ==============================================
# 📄 CONFIGURACIÓN DE LOGGING GLOBAL
# ==============================================

logging.basicConfig(
    level=logging.WARNING,  # Solo mostrar WARNING y superiores
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)

# Reducir verbosidad de dependencias
logging.getLogger('werkzeug').setLevel(logging.ERROR)
logging.getLogger('pymongo').setLevel(logging.ERROR)
logging.getLogger('urllib3').setLevel(logging.ERROR)
logging.getLogger('boto3').setLevel(logging.ERROR)
logging.getLogger('botocore').setLevel(logging.ERROR)

# Manejo de excepciones no capturadas
def handle_exception(exc_type, exc_value, exc_traceback):
    if not issubclass(exc_type, KeyboardInterrupt):
        logging.error("Excepción no capturada", exc_info=(exc_type, exc_value, exc_traceback))

sys.excepthook = handle_exception

# ==============================================
# 📄 FUNCIONES PARA SUBIDA Y GESTIÓN DE ARCHIVOS EN S3
# ==============================================

def upload_file_to_s3(file_path, object_name=None, max_retries=3, delete_local=True):
    """Sube un archivo a S3, verifica subida y elimina el local si éxito"""
    if object_name is None:
        object_name = os.path.basename(file_path)

    if not os.path.exists(file_path):
        app.logger.error(f"El archivo {file_path} no existe.")
        return False

    successful = False
    attempt = 0
    while attempt < max_retries and not successful:
        attempt += 1
        try:
            current_app.s3_client.upload_file(file_path, current_app.S3_BUCKET_NAME, object_name)
            current_app.s3_client.head_object(Bucket=current_app.S3_BUCKET_NAME, Key=object_name)
            successful = True
            app.logger.info(f"✅ Archivo {object_name} subido a S3 exitosamente")
        except Exception as e:
            app.logger.error(f"❌ Intento {attempt}: Error al subir {object_name}: {str(e)}")
            if attempt < max_retries:
                time.sleep(2 ** attempt)

    if successful and delete_local:
        try:
            os.remove(file_path)
            app.logger.info(f"🗑️ Archivo local eliminado: {file_path}")
        except Exception as e:
            app.logger.warning(f"⚠️ No se pudo eliminar {file_path}: {e}")

    return successful

def delete_file_from_s3(object_name):
    """Elimina un archivo de un bucket S3"""
    try:
        current_app.s3_client.delete_object(Bucket=current_app.S3_BUCKET_NAME, Key=object_name)
        app.logger.info(f"✅ Archivo eliminado de S3: {object_name}")
        return True
    except Exception as e:
        app.logger.error(f"❌ Error eliminando {object_name} de S3: {str(e)}")
        return False

def get_s3_url(object_name, expiration=3600):
    """Genera una URL firmada para acceder temporalmente a un objeto en S3"""
    try:
        url = current_app.s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': current_app.S3_BUCKET_NAME, 'Key': object_name},
            ExpiresIn=expiration
        )
        return url
    except Exception as e:
        app.logger.error(f"Error generando URL firmada para {object_name}: {e}")
        return None

# ==============================================
# 🔐 RUTAS DE AUTENTICACIÓN
# ==============================================

from werkzeug.security import generate_password_hash, check_password_hash
import secrets

def verify_scrypt_password(stored_hash, provided_password):
    """
    Verifica una contraseña usando el hash scrypt
    """
    try:
        # Extraer los parámetros del hash almacenado
        parts = stored_hash.split('$')
        if len(parts) != 4:
            return False
        salt = parts[2]
        stored_hash_val = parts[3]
        # Generar el hash de la contraseña proporcionada usando hashlib.scrypt
        new_hash = hashlib.scrypt(provided_password.encode('utf-8'), salt=salt.encode('utf-8'), n=32768, r=8, p=1, dklen=64)
        # Comparar los hashes
        return stored_hash_val == new_hash.hex()
    except Exception as e:
        print(f"Error verificando contraseña: {str(e)}")
        return False

# ==============================================
# 📋 RUTAS PRINCIPALES (Tablas, Catálogo, Archivos)
# ==============================================

# --- Home: Decide a dónde enviar al usuario ---
@app.route("/")
def home():
    if not (session.get('username') or session.get('email')):
        return render_template("welcome.html")
    if "selected_table" in session:
        return redirect(url_for("catalog"))
    else:
        return redirect(url_for("tables"))

# --- Página de bienvenida ---
@app.route("/welcome")
def welcome():
    return render_template("welcome.html", css_version=CSS_VERSION)

# --- Seleccionar una tabla ---
@app.route("/select_table/<table_id>")
@login_required
def select_table(table_id):
    owner = session.get('username') or session.get('email')
    table = current_app.spreadsheets_collection.find_one({"_id": ObjectId(table_id)})
    if not table:
        flash("Tabla no encontrada.", "error")
        return redirect(url_for("tables"))
    session["selected_table"] = table["filename"]
    session["selected_table_id"] = str(table["_id"])
    session["selected_table_name"] = table["name"]
    return redirect(url_for("catalog"))

# --- Catálogo: Mostrar los registros de una tabla ---
@app.route("/catalog", methods=["GET", "POST"])
def catalog():
    if "username" not in session:
        return redirect(url_for("welcome"))
    if "selected_table" not in session:
        flash("Por favor, selecciona una tabla.", "warning")
        return redirect(url_for("tables"))
    selected_table = session["selected_table"]
    table_info = current_app.spreadsheets_collection.find_one({"filename": selected_table})
    if not table_info:
        flash("La tabla seleccionada no existe.", "error")
        return redirect(url_for("tables"))
    # Control de acceso: solo admin o dueño puede ver
    # Verificación de rol desactivada)
    headers = table_info.get("headers", [])

    # Mostrar registros ordenados por número
    pipeline = [
        {"$match": {"table": selected_table}},
        {"$addFields": {"NumeroOrdenacion": {"$toInt": {"$ifNull": [{"$toInt": "$Número"}, "$Número"]}}}},
        {"$sort": {"NumeroOrdenacion": 1}}
    ]
    registros = list(current_app.catalog_collection.aggregate(pipeline))

    # Método POST: Insertar nuevo registro
    if request.method == "POST":
        form_data = {k.strip(): v.strip() for k, v in request.form.items()}

        id_field = headers[0]  # Normalmente "Número" o similar

        if not form_data.get(id_field):
            flash(f"Error: El campo {id_field} es obligatorio.", "error")
            return render_template("index.html", data=registros, headers=headers)

        if any(item.get(id_field) == form_data[id_field] for item in registros):
            flash(f"Error: Ya existe un registro con {id_field} {form_data[id_field]}.", "error")
            return render_template("index.html", data=registros, headers=headers)

        nuevo_registro = {"Número": len(registros) + 1, "table": selected_table}
        
        for header in headers:
            safe_header = header.replace(" ", "_").replace(".", "_")
            nuevo_registro[safe_header] = form_data.get(header, "")

        # Manejo de imágenes (se suben a S3)
        files = request.files.getlist("imagenes")
        rutas_imagenes = []
        for file in files[:3]:
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                unique_filename = f"{timestamp}_{secrets.token_hex(4)}_{filename}"
                temp_path = os.path.join(app.config["UPLOAD_FOLDER"], unique_filename)
                file.save(temp_path)

                if upload_file_to_s3(temp_path, unique_filename):
                    rutas_imagenes.append(f"s3://{current_app.S3_BUCKET_NAME}/{unique_filename}")

        nuevo_registro["Imagenes"] = rutas_imagenes

        current_app.catalog_collection.insert_one(nuevo_registro)
        return redirect(url_for("catalog"))

    # Método GET: Mostrar registros
    return render_template("index.html", data=registros, headers=headers)

# ==============================================
# ✏️ RUTAS PARA EDITAR Y ELIMINAR REGISTROS
# ==============================================

# --- Editar un registro ---
@app.route("/editar/<id>", methods=["GET", "POST"])
def editar(id):
    if "username" not in session:
        return redirect(url_for("auth.login"))

    if "selected_table" not in session:
        flash("Selecciona una tabla primero.", "warning")
        return redirect(url_for("tables"))

    selected_table = session["selected_table"]
    table_info = current_app.spreadsheets_collection.find_one({"filename": selected_table})

    if not table_info:
        flash("Tabla no encontrada.", "error")
        return redirect(url_for("tables"))

    headers = table_info.get("headers", [])
    id_field = headers[0]
    safe_id_field = id_field.replace(" ", "_").replace(".", "_")

    registro = current_app.catalog_collection.find_one({"_id": ObjectId(id)})

    if not registro:
        flash("Registro no encontrado.", "error")
        return redirect(url_for("catalog"))

    # --- GET: Mostrar el formulario de edición ---
    if request.method == "GET":
        headers_form = [h for h in headers if h != "Imagenes"]
        return render_template("editar.html", 
                               registro=registro,
                               headers=headers_form,
                               imagenes_actuales=registro.get("Imagenes", [None, None, None]))

    # --- POST: Guardar cambios ---
    update_data = {
        "Número": registro["Número"],
        "table": selected_table
    }

    for header in headers:
        if header != "Imagenes" and header != "Número":
            safe_header = header.replace(" ", "_").replace(".", "_")
            update_data[safe_header] = request.form.get(header, "").strip()

    # Manejo de nuevas imágenes
    rutas_imagenes = registro.get("Imagenes", [None, None, None])
    
    for idx, field_name in enumerate(["imagen1", "imagen2", "imagen3"]):
        imagen = request.files.get(field_name)
        if imagen and imagen.filename and allowed_file(imagen.filename):
            filename = secure_filename(imagen.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            unique_filename = f"{timestamp}_{secrets.token_hex(4)}_{filename}"
            temp_path = os.path.join(app.config["UPLOAD_FOLDER"], unique_filename)
            imagen.save(temp_path)

            if upload_file_to_s3(temp_path, unique_filename):
                rutas_imagenes[idx] = f"s3://{current_app.S3_BUCKET_NAME}/{unique_filename}"

    # Manejar eliminación de imágenes
    for idx in range(3):
        if request.form.get(f"remove_img{idx+1}") == "on":
            eliminar_archivo_imagen(rutas_imagenes[idx])
            rutas_imagenes[idx] = None

    update_data["Imagenes"] = rutas_imagenes

    # Actualizar el registro en MongoDB
    current_app.catalog_collection.update_one(
        {"_id": ObjectId(id)},
        {"$set": update_data}
    )

    flash("Registro actualizado correctamente.", "success")
    return redirect(url_for("catalog"))

# --- Eliminar una tabla completa ---
@app.route("/delete_table/<table_id>", methods=["POST"])
def delete_table(table_id):
    if "username" not in session:
        return redirect(url_for("auth.login"))

    table = current_app.spreadsheets_collection.find_one({"_id": ObjectId(table_id), "owner": session["username"]})

    if not table:
        flash("Tabla no encontrada o no tienes permiso.", "error")
        return redirect(url_for("tables"))

    filepath = os.path.join(SPREADSHEET_FOLDER, table["filename"])

    if os.path.exists(filepath):
        os.remove(filepath)

    current_app.spreadsheets_collection.delete_one({"_id": ObjectId(table_id)})

    if session.get("selected_table") == table["filename"]:
        session.pop("selected_table", None)

    flash("Tabla eliminada exitosamente.", "success")
    return redirect(url_for("tables"))

# ==============================================
# 📦 RUTAS PARA DESCARGAR CATÁLOGO COMO ZIP
# ==============================================

import tempfile
import zipfile

@app.route("/descargar-excel")
def descargar_excel():
    if "username" not in session:
        return redirect(url_for("auth.login"))

    spreadsheet_path = get_current_spreadsheet()

    if not spreadsheet_path or not os.path.exists(spreadsheet_path):
        return "❌ El Excel no existe aún."

    # Crear un archivo temporal ZIP
    temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    with zipfile.ZipFile(temp_zip.name, "w") as zf:
        # Incluir el Excel
        zf.write(spreadsheet_path, arcname=os.path.basename(spreadsheet_path))

        # Incluir imágenes (locales o de S3)
        data = leer_datos_excel(spreadsheet_path)
        image_paths = set()

        for row in data:
            for ruta in row.get("Imagenes", []):
                if ruta:
                    if ruta.startswith('s3://'):
                        try:
                            s3_parts = ruta[5:].split('/', 1)
                            if len(s3_parts) == 2:
                                bucket_name, object_key = s3_parts
                                if bucket_name == current_app.S3_BUCKET_NAME:
                                    temp_path = os.path.join(tempfile.gettempdir(), os.path.basename(object_key))
                                    current_app.s3_client.download_file(bucket_name, object_key, temp_path)
                                    image_paths.add(temp_path)
                        except Exception as e:
                            print(f"Error al descargar imagen S3: {e}")
                    else:
                        local_path = os.path.join(app.root_path, ruta.lstrip("/"))
                        if os.path.exists(local_path):
                            image_paths.add(local_path)

        for img_path in image_paths:
            zf.write(img_path, arcname=f"imagenes/{os.path.basename(img_path)}")

    return send_from_directory(directory=os.path.dirname(temp_zip.name),
                               path=os.path.basename(temp_zip.name),
                               as_attachment=True,
                               download_name="catalogo.zip")

# ==============================================
# 🖼️ RUTA PARA SERVIR IMÁGENES SUBIDAS
# ==============================================


# ==============================================
# 🔢 RENOMBRAR REGISTROS AUTOMÁTICAMENTE
# ==============================================

def renumerar_registros(table_name):
    """Renumera todos los registros de una tabla específica en orden secuencial"""
    registros = list(current_app.catalog_collection.find({"table": table_name}).sort("Número", 1))

    for i, registro in enumerate(registros, 1):
        if registro.get("Número") != i:
            current_app.catalog_collection.update_one(
                {"_id": registro["_id"]},
                {"$set": {"Número": i}}
            )
    
    return len(registros)

@app.route("/renumerar/<table_name>")
def renumerar(table_name):
    """Renumera todos los registros de una tabla y redirecciona al catálogo"""
    if "username" not in session:
        return redirect(url_for("welcome"))

    try:
        total = renumerar_registros(table_name)
        flash(f"✅ Se renumeraron {total} registros.", "success")
    except Exception as e:
        flash(f"❌ Error al renumerar registros: {e}", "error")

    session["selected_table"] = table_name
    return redirect(url_for("catalog"))

# ==============================================
# ❌ MANEJO DE ERRORES PERSONALIZADOS
# ==============================================

@app.errorhandler(404)
def page_not_found(e):
    """Manejo de error 404: Página no encontrada"""
    return render_template('not_found.html'), 404

@app.errorhandler(500)
def server_error(e):
    """Manejo de error 500: Error interno del servidor"""
    return render_template('error.html', error=str(e)), 500

# ==============================================
# 🎨 RUTA PARA PROBAR LOS ESTILOS VISUALES
# ==============================================

@app.route('/test_styles')
def test_styles():
    """Ruta para probar el correcto funcionamiento de los CSS y otros recursos estáticos"""
    return render_template('test_styles.html')

# ==============================================
# 🛠️ GESTOR DE SCRIPTS
# ==============================================

# Estas rutas han sido trasladadas al blueprint scripts_bp en scripts_routes.py
# con el prefijo /admin/tools

# @app.route('/tools')
# @admin_required
# def tools_dashboard():
#     """Ruta para acceder al gestor de scripts del sistema"""
#     import os
#     import glob
#     
#     # Definir las categorías de scripts
#     script_categories = {
#         'maintenance': {
#             'name': 'Mantenimiento',
#             'description': 'Scripts para iniciar, supervisar y mantener la aplicación',
#             'path': os.path.join(ROOT_DIR, 'scripts/maintenance')
#         },
#         'organization': {
#             'name': 'Organización',
#             'description': 'Scripts para organizar y limpiar la estructura del proyecto',
#             'path': ROOT_DIR
#         },
#         'deployment': {
#             'name': 'Despliegue',
#             'description': 'Scripts relacionados con el despliegue de la aplicación',
#             'path': os.path.join(ROOT_DIR, 'scripts/deployment')
#         },
#         'system': {
#             'name': 'Sistema',
#             'description': 'Scripts para tareas del sistema operativo y configuración',
#             'path': os.path.join(ROOT_DIR, 'scripts/system')
#         }
#     }
#     
#     # Recopilar todos los scripts
#     scripts = []
#     
#     # Scripts de mantenimiento
#     maintenance_scripts = glob.glob(os.path.join(script_categories['maintenance']['path'], '*.sh'))
#     for script_path in maintenance_scripts:
#         scripts.append({
#             'name': os.path.basename(script_path),
#             'path': script_path,
#             'category': 'maintenance',
#             'description': 'Script de mantenimiento del sistema'
#         })
#     
#     # Scripts de organización
#     organization_scripts = [
#         os.path.join(ROOT_DIR, 'cleanup_duplicates.sh'),
#         os.path.join(ROOT_DIR, 'organize_root_scripts.sh'),
#         os.path.join(ROOT_DIR, 'reorganize_project.sh')
#     ]
#     for script_path in organization_scripts:
#         if os.path.exists(script_path):
#             scripts.append({
#                 'name': os.path.basename(script_path),
#                 'path': script_path,
#                 'category': 'organization',
#                 'description': 'Script de organización del proyecto'
#             })
#     
#     # Otros scripts en el directorio raíz
#     root_scripts = glob.glob(os.path.join(ROOT_DIR, '*.sh'))
#     for script_path in root_scripts:
#         if script_path not in organization_scripts:
#             scripts.append({
#                 'name': os.path.basename(script_path),
#                 'path': script_path,
#                 'category': 'system',
#                 'description': 'Script del sistema'
#             })
#     
#     return render_template('tools_dashboard.html', scripts=scripts, categories=script_categories)

# @app.route('/tools/view_script')
# @admin_required
# def view_script():
#     """Ruta para ver el contenido de un script"""
#     script_path = request.args.get('path')
#     
#     if not script_path:
#         return jsonify({'error': 'No se ha especificado la ruta del script'}), 400
    
#     # Verificar que el archivo está dentro del directorio del proyecto
#     if not script_path.startswith(ROOT_DIR):
#         return jsonify({'error': 'No se permite acceder a archivos fuera del directorio del proyecto'}), 403
#     
#     try:
#         with open(script_path, 'r') as file:
#             content = file.read()
#         return jsonify({'content': content, 'name': os.path.basename(script_path)})
#     except Exception as e:
#         return jsonify({'error': f'Error al leer el archivo: {str(e)}'}), 500

# @app.route('/tools/run_script')
# @admin_required
# def run_script():
#     """Ruta para ejecutar un script"""
#     script_path = request.args.get('path')
#     
#     if not script_path:
#         return jsonify({'error': 'No se ha especificado la ruta del script'}), 400
#     
#     # Verificar que la ruta es válida y apunta a un archivo .sh
#     if not os.path.exists(script_path) or not os.path.isfile(script_path) or not script_path.endswith('.sh'):
#         return jsonify({'error': 'El archivo no existe o no es un script .sh válido'}), 404
#     
#     # Verificar que el archivo está dentro del directorio del proyecto
#     if not script_path.startswith(ROOT_DIR):
#         return jsonify({'error': 'No se permite ejecutar archivos fuera del directorio del proyecto'}), 403
#     
#     # Verificar que el archivo tiene permisos de ejecución
#     if not os.access(script_path, os.X_OK):
#         try:
#             os.chmod(script_path, 0o755)
#         except Exception as e:
#             return jsonify({'error': f'No se pudieron establecer permisos de ejecución: {str(e)}'}), 500
#     
#     # Ejecutar el script y capturar la salida
#     try:
#         import subprocess
#         process = subprocess.Popen(
#             f"cd {os.path.dirname(script_path)} && {script_path}",
#             shell=True,
#             stdout=subprocess.PIPE,
#             stderr=subprocess.PIPE,
#             text=True
#         )
#         stdout, stderr = process.communicate(timeout=30)  # Timeout de 30 segundos
#         
#         result = {
#             'script': os.path.basename(script_path),
#             'exit_code': process.returncode,
#             'output': stdout,
#             'error': stderr,
#             'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#         }
#         
#         return jsonify(result)
#     except subprocess.TimeoutExpired:
#         return jsonify({
#             'script': os.path.basename(script_path),
#             'error': 'El script tardó demasiado tiempo en ejecutarse (más de 30 segundos)',
#             'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#         }), 408
#     except Exception as e:
#         return jsonify({
#             'script': os.path.basename(script_path),
#             'error': f'Error al ejecutar el script: {str(e)}',
#             'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#         }), 500

# ==============================================
# 🔥 AJUSTE FINAL: LANZAMIENTO DE LA APP
# ==============================================

# if __name__ == '__main__':
#     app.run(debug=False, host='0.0.0.0')

# ==============================================
# 📄 RUTAS PARA DASHBOARDS
# ==============================================

# --- Dashboard para usuario normal ---
# (Eliminado: esta ruta ahora está en el blueprint main_bp)

# ==============================================
# 🔐 RUTAS PROTEGIDAS
# ==============================================

# --- Decorador para rutas protegidas solo admin ---
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "username" not in session or session.get("role") != "admin":
            flash("No tiene permisos para acceder a esta página", "error")
            return redirect(url_for("auth.login"))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/ping')
def ping():
    return 'pong'

@app.route('/editar_fila/<fila_id>', methods=['GET', 'POST'])
@login_required
def editar_fila(fila_id):
    owner = session["username"]
    fila = current_app.catalog_collection.find_one({'_id': ObjectId(fila_id)})
    if not fila:
        flash('Fila no encontrada.', 'error')
        return redirect(url_for('tables'))
    # Obtener info de la tabla
    table_info = current_app.spreadsheets_collection.find_one({'filename': fila['table']})
    if not table_info:
        flash('Tabla asociada no encontrada.', 'error')
        return redirect(url_for('tables'))
    # Control de acceso
    # Verificación de rol desactivada))
    headers = table_info.get('headers', [])
    if request.method == 'POST':
        update_data = {}
        for header in headers:
            if header != 'Número' and header != 'Imagenes':
                update_data[header] = request.form.get(header, '').strip()
        current_app.catalog_collection.update_one({'_id': ObjectId(fila_id)}, {'$set': update_data})
        flash('Fila actualizada correctamente.', 'success')
        return redirect(url_for('ver_tabla', table_id=str(table_info['_id'])))
    return render_template('editar_fila.html', fila=fila, headers=headers, catalog=table_info)

@app.route('/test_session')
def test_session():
    from flask import session
    session['test'] = 'ok'
    return f"Valor de session['test']: {session.get('test')}"

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        usuario = current_app.users_collection.find_one({'username': username})
        if usuario and check_password_hash(usuario['password'], password):
            session['username'] = usuario['username']
            session['user_id'] = str(usuario['_id'])
            session['role'] = usuario.get('role', 'user')
            session.permanent = True
            return redirect(url_for('home'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    return render_template('login.html')

