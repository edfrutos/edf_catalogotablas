#!/usr/bin/env python3
from openpyxl import load_workbook
import os

filepath = "/spreadsheets/table_7e08c6bfa70080ae.xlsx"

print(f"\nVerificando archivo: {os.path.basename(filepath)}")
try:
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    print("\nContenido del archivo:")
    for row in ws.iter_rows(values_only=True):
        print(row)
    wb.close()
except Exception as e:
    print(f"Error al leer el archivo: {str(e)}")
