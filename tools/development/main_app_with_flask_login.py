# ==============================================
# 📄 CONFIGURACIONES INICIALES
# ==============================================

# --- Imports estándar ---
# traceback removido - no usado
import hashlib
import logging
import os
import secrets
import sys
import tempfile
import time
import zipfile
from datetime import datetime
from typing import Any

# --- Imports de terceros ---
import openpyxl  # type: ignore
from bson import ObjectId
from dotenv import load_dotenv
from flask import (
    Flask,
    abort,
    current_app,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_from_directory,
    session,
    url_for,
)

# Importar Flask-Login y modelo User para configuración
from flask_login import LoginManager
from werkzeug.security import check_password_hash
from werkzeug.utils import secure_filename

from app.decorators import login_required  # type: ignore

# Importar otros módulos de utilidad
from app.extensions import init_extensions

# Configurar logging unificado desde el inicio
from app.logging_unified import setup_unified_logging
from app.models.user import User
from flask_session import Session  # type: ignore


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in {
        "png",
        "jpg",
        "jpeg",
        "gif",
    }


def eliminar_archivo_imagen(ruta: str) -> None:
    if not ruta:
        return
    if ruta.startswith("s3://"):
        s3_parts = ruta[5:].split("/", 1)
        if len(s3_parts) == 2:
            bucket_name, object_key = s3_parts
            if bucket_name == current_app.config["S3_BUCKET_NAME"]:  # type: ignore
                try:
                    current_app.s3_client.delete_object(  # type: ignore
                        Bucket=bucket_name, Key=object_key
                    )
                except Exception as e:
                    current_app.logger.error(f"Error eliminando imagen de S3: {e}")
    else:
        local_path = os.path.join(
            current_app.config["UPLOAD_FOLDER"], os.path.basename(ruta)
        )
        if os.path.exists(local_path):
            try:
                os.remove(local_path)
            except Exception as e:
                current_app.logger.error(f"Error eliminando imagen local: {e}")


def get_current_spreadsheet() -> str | None:
    selected_table = session.get("selected_table")
    if not selected_table:
        return None
    return os.path.join(current_app.config["UPLOAD_FOLDER"], selected_table)


def leer_datos_excel(filepath: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(filepath)
    hoja = wb.active
    if hoja is None:
        wb.close()
        return []
    headers = [
        cell.value for cell in next(hoja.iter_rows(min_row=1, max_row=1))
    ]  # type: ignore
    data = []
    for row in hoja.iter_rows(min_row=2, values_only=True):  # type: ignore
        data.append(dict(zip(headers, row)))
    wb.close()
    return data


# ==============================================
# 📄 CONFIGURACIÓN INICIAL DE FLASK
# ==============================================

# --- Cargar variables de entorno ---
load_dotenv()


def create_app():
    print("DEBUG: create_app ejecutado")
    """Crea y configura la aplicación Flask"""
    ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
    TEMPLATE_DIR = os.path.join(ROOT_DIR, "app", "templates")
    STATIC_DIR = os.path.join(ROOT_DIR, "app", "static")
    app = Flask(
        __name__,
        template_folder=TEMPLATE_DIR,
        static_folder=STATIC_DIR,
        static_url_path="/static",
    )

    # --- Registrar blueprint de imágenes ---
    from app.routes.images_routes import images_bp

    app.register_blueprint(images_bp)

    # Cargar configuración apropiada según el entorno
    if getattr(sys, "frozen", False):
        # Aplicación empaquetada - intentar configuración embebida
        try:
            from app.config_embedded import EmbeddedConfig  # type: ignore

            app.config.from_object(EmbeddedConfig)
        except ImportError:
            # Fallback a configuración normal si no existe embebida
            app.config.from_object("config.Config")
    else:
        # Desarrollo - usar configuración normal
        app.config.from_object("config.Config")

    # CONFIGURACIÓN DE SESIÓN UNIFICADA
    # La configuración de sesiones ahora se maneja centralmente en app/config.py
    # Asegurar que el directorio de sesiones existe
    session_dir = app.config.get("SESSION_FILE_DIR")
    if session_dir:
        if not os.path.isabs(session_dir):
            session_dir = os.path.join(ROOT_DIR, session_dir)
            app.config["SESSION_FILE_DIR"] = session_dir
        os.makedirs(session_dir, exist_ok=True)
        app.logger.info(f"✅ Directorio de sesiones configurado: {session_dir}")

    app.logger.info("✅ Configuración de sesión unificada aplicada desde config.py")

    # Inicializar Flask-Session (debe estar después de configurar app.config)
    Session(app)
    app.logger.info("✅ Flask-Session inicializado correctamente")

    # Inicializar sistema de logging unificado
    setup_unified_logging(app)

    # Inicializar extensiones (Flask-Login, Flask-Mail, etc.)
    init_extensions(app)
    app.logger.info(
        "✅ Extensiones inicializadas correctamente (Flask-Login, Flask-Mail, etc.)"
    )

    # Configurar sesión permanente por defecto
    @app.before_request
    def make_session_permanent():
        session.permanent = True

    # Inicializar la conexión global a MongoDB (para funciones legacy y modelos)
    try:
        from app.database import get_mongo_client, get_mongo_db, initialize_db

        initialize_db(app)
        app.logger.info("✅ Conexión global a MongoDB inicializada (initialize_db)")
        # Refuerza la asignación de app.db y colecciones SIEMPRE tras inicializar
        client = get_mongo_client()
        db = get_mongo_db()
        app.mongo_client = client  # type: ignore
        app.db = db  # type: ignore
        if db is not None:
            app.users_collection = db["users"]  # type: ignore
            app.resets_collection = db["password_resets"]  # type: ignore
            app.catalog_collection = db["67b8c24a7fdc72dd4d8703cf"]  # type: ignore
            app.spreadsheets_collection = db["spreadsheets"]  # type: ignore
        else:
            app.users_collection = None  # type: ignore
            app.resets_collection = None  # type: ignore
            app.catalog_collection = None  # type: ignore
            app.spreadsheets_collection = None  # type: ignore
    except Exception as e:
        app.logger.error(f"❌ Error inicializando la conexión global a MongoDB: {e}")
        app.db = None  # type: ignore
        app.users_collection = None  # type: ignore
        app.resets_collection = None  # type: ignore
        app.catalog_collection = None  # type: ignore
        app.spreadsheets_collection = None  # type: ignore

    # =================== FUNCIONES AUXILIARES ===================
    # (Eliminar las definiciones internas de eliminar_archivo_imagen, get_current_spreadsheet y leer_datos_excel aquí)

    # =================== RUTAS Y HANDLERS ===================
    # (Las rutas pueden ahora usar current_app.users_collection, etc.)
    # Ejemplo de uso en una ruta:
    # usuario = current_app.users_collection.find_one({...})
    # s3 = current_app.s3_client
    # bucket = current_app.S3_BUCKET_NAME
    # ...

    # --- Versión CSS automática para evitar caché ---
    CSS_VERSION = datetime.now().strftime("%Y%m%d%H%M%S")

    # --- Función para saber si un valor es flotante ---
    def is_float(value):
        try:
            float(value)
            return True
        except (ValueError, TypeError):
            return False

    # --- Inyectar variables globales en todas las plantillas ---
    @app.context_processor
    def inject_template_vars():
        return {"now": datetime.now(), "css_version": CSS_VERSION}

    # --- Configurar archivos estáticos personalizados ---
    @app.route("/static/<path:filename>")
    def custom_static(filename):
        static_dir = os.path.join(ROOT_DIR, "app", "static")
        try:
            app.logger.info(
                f"Intentando servir archivo estático: {filename} desde {static_dir}"
            )

            # Verificar que el archivo existe
            full_path = os.path.join(static_dir, filename)
            if not os.path.exists(full_path):
                app.logger.error(f"Archivo no encontrado: {full_path}")
                abort(404)

            return send_from_directory(static_dir, filename)
        except Exception as e:
            app.logger.error(f"Error sirviendo archivo estático {filename}: {str(e)}")
            app.logger.error(f"Ruta completa: {os.path.join(static_dir, filename)}")
            abort(404)

    # --- Decorador para rutas protegidas ---
    # def login_required(f):
    #     def wrapper(*args, **kwargs):
    #         if 'user_id' not in session:
    #             flash('Debes iniciar sesión primero.', 'warning')
    #             return redirect(url_for('auth.login'))
    #         return f(*args, **kwargs)
    #     wrapper.__name__ = f.__name__
    #     return wrapper

    # Registrar blueprints principales
    from app.routes.admin_routes import (  # Importar también admin_logs_bp
        admin_bp,
        admin_logs_bp,
    )
    from app.routes.catalog_images_routes import image_bp
    from app.routes.catalogs_routes import catalogs_bp
    from app.routes.dev_template import (  # Blueprint para plantilla de desarrollo
        bp_dev_template,
    )
    from app.routes.emergency_access import emergency_bp
    from app.routes.error_routes import errors_bp
    from app.routes.main_routes import main_bp
    from app.routes.maintenance_routes import (  # Blueprint para mantenimiento
        maintenance_bp,
    )

    # Debug blueprints ya registrados en otras partes
    from app.routes.scripts_routes import (  # Blueprint para gestión de scripts
        scripts_bp,
    )
    from app.routes.usuarios_routes import usuarios_bp

    print("ANTES DE BLUEPRINTS", app.db)  # type: ignore
    app.register_blueprint(main_bp)
    app.register_blueprint(catalogs_bp)
    app.register_blueprint(image_bp)
    app.register_blueprint(usuarios_bp)
    app.register_blueprint(admin_bp, url_prefix="/admin")
    app.register_blueprint(
        admin_logs_bp, url_prefix="/admin"
    )  # Registrar admin_logs_bp
    app.register_blueprint(
        scripts_bp
    )  # Blueprint para gestión de scripts (/admin/tools)
    app.register_blueprint(
        maintenance_bp, url_prefix="/admin"
    )  # Blueprint para mantenimiento
    app.register_blueprint(bp_dev_template)  # Blueprint para plantilla de desarrollo
    app.register_blueprint(errors_bp)
    app.register_blueprint(emergency_bp)  # <-- REGISTRO DE EMERGENCIA
    print("DESPUÉS DE BLUEPRINTS", app.db)  # type: ignore

    # Inicializar Flask-Login
    login_manager = LoginManager()
    # Para evitar error de tipo, asignar login_view con setattr
    login_manager.login_view = "auth.login"
    login_manager.init_app(app)

    # Definir user_loader para Flask-Login
    @login_manager.user_loader
    def load_user(user_id):
        from bson import ObjectId

        try:
            # Acceder a users_collection desde g para evitar error de tipo
            from flask import g

            user_data = (
                g.db["users"].find_one({"_id": ObjectId(user_id)}) if g.db else None
            )
            if user_data:
                return User(user_data)
        except Exception as e:
            app.logger.error(f"Error cargando usuario para Flask-Login: {e}")
        return None

    return app


app = create_app()

if __name__ == "__main__":
    port = 5002
    if len(sys.argv) > 1 and sys.argv[1] == "--port" and len(sys.argv) > 2:
        try:
            port = int(sys.argv[2])
        except ValueError:
            port = 5002
    app.run(debug=False, host="0.0.0.0", port=port)
