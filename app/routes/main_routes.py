# app/routes/main_routes.py

import csv
import logging
import os
import secrets
import uuid
from datetime import datetime

import openpyxl
from bson.objectid import ObjectId
from flask import (
    Blueprint,
    Response,
    current_app,
    flash,
    g,
    redirect,
    render_template,
    request,
    session,
    url_for,
)
from openpyxl import Workbook
from werkzeug.routing import BuildError
from werkzeug.utils import secure_filename

from app import notifications
from app.database import get_mongo_db
from app.decorators import login_required
from app.utils.image_utils import get_images_for_template


def get_upload_dir():
    """Obtener directorio de uploads"""
    static_folder = current_app.static_folder or "static"
    folder = os.path.join(static_folder, "uploads")
    os.makedirs(folder, exist_ok=True)
    return folder


main_bp = Blueprint("main", __name__)
logger = logging.getLogger(__name__)


@main_bp.route("/")
def index():
    # Redirigir a welcome si no est√° logueado
    if "user_id" not in session:
        return redirect(url_for("main.welcome"))
    return redirect(url_for("main.dashboard_user"))


@main_bp.route("/clear-flash-messages", methods=["POST"])
def clear_flash_messages():
    """Limpiar mensajes flash de la sesi√≥n"""
    from flask import session

    session.pop("_flashes", None)
    return {"status": "success", "message": "Mensajes flash limpiados"}


@main_bp.route("/welcome")
def welcome():
    return render_template("welcome.html")


@main_bp.route("/contacto", methods=["GET", "POST"])
def contacto():
    if request.method == "POST":
        # Obtener datos del formulario
        nombre = request.form.get("nombre", "").strip()
        email = request.form.get("email", "").strip()
        asunto = request.form.get("asunto", "").strip()
        mensaje = request.form.get("mensaje", "").strip()

        # Validaci√≥n b√°sica
        if not all([nombre, email, asunto, mensaje]):
            flash("Todos los campos son obligatorios.", "error")
            return render_template("contacto.html")

        # Enviar email real usando el sistema de notificaciones
        subject = f"[Contacto] {asunto}"
        body_html = f"""
        <h3>Nuevo mensaje de contacto recibido</h3>
        <ul>
            <li><strong>Nombre:</strong> {nombre}</li>
            <li><strong>Email:</strong> {email}</li>
            <li><strong>Asunto:</strong> {asunto}</li>
        </ul>
        <p><strong>Mensaje:</strong></p>
        <div style='white-space: pre-line; border:1px solid #eee;
             padding:10px; background:#f9f9f9;'>{mensaje}</div>
        <hr>
        <small>Este mensaje fue enviado desde el formulario de contacto de la aplicaci√≥n.</small>
        """

        # Destinatario: edfrutos@gmail.com
        recipients = ["edfrutos@gmail.com"]

        try:
            from app import notifications

            ok = notifications.send_email(subject, body_html, recipients)
            if ok:
                flash(
                    f"¬°Gracias {nombre}! Tu mensaje ha sido enviado correctamente. "
                    + "Te contactaremos pronto.",
                    "success",
                )
            else:
                flash(
                    "Tu mensaje se ha recibido pero hubo un problema al enviarlo. "
                    + "Te contactaremos pronto.",
                    "warning",
                )
        except Exception as e:
            logger.error(f"Error enviando email de contacto: {str(e)}")
            flash(
                "Tu mensaje se ha recibido pero hubo un problema t√©cnico. "
                + "Te contactaremos pronto.",
                "warning",
            )

        logger.info(
            f"Formulario de contacto enviado - Nombre: {nombre}, "
            + f"Email: {email}, Asunto: {asunto}"
        )

        # Redirigir para evitar reenv√≠o del formulario
        return redirect(url_for("main.contacto"))

    # GET request - mostrar formulario
    return render_template("contacto.html")


@main_bp.route("/dashboard_user")
def dashboard_user():
    # Protecci√≥n: requerir login
    if not session.get("username") and not session.get("user_id"):
        flash("Debe iniciar sesi√≥n para acceder al dashboard de usuario", "warning")
        return redirect(url_for("auth.login", next=request.url))
    if session.get("role") == "admin":
        return redirect(url_for("admin.dashboard_admin"))

    # =========================================================================
    # VERIFICACI√ìN CR√çTICA: DETECTAR CONTRASE√ëA TEMPORAL
    # =========================================================================
    user_id = session.get("user_id")
    if user_id:
        try:
            from bson import ObjectId

            from app.models.database import get_users_collection

            users_collection = get_users_collection()
            user = users_collection.find_one({"_id": ObjectId(user_id)})

            if user:
                # Verificar si tiene contrase√±a temporal
                has_temp_password = (
                    user.get("temp_password", False)
                    or user.get("must_change_password", False)
                    or user.get("password_reset_required", False)
                )

                if has_temp_password:
                    # Usuario con contrase√±a temporal - redirigir INMEDIATAMENTE
                    session["temp_reset_user_id"] = str(user.get("_id"))
                    session["temp_reset_email"] = user.get("email", "")
                    session["temp_reset_username"] = user.get("username", "")

                    flash(
                        "ATENCI√ìN: Tu cuenta usa una contrase√±a temporal. Debes crear una nueva contrase√±a para usar el sistema completo.",
                        "error",
                    )
                    return redirect(url_for("auth.temp_password_reset"))

        except Exception as e:
            print(f"[ERROR] Error verificando contrase√±a temporal: {e}")
            # Continuar normal si hay error, no bloquear el acceso
    # Acceso solo a datos propios
    username = session.get("username")
    email = session.get("email")
    nombre = session.get("nombre", username)
    posibles = set([username, email, nombre])
    posibles = {v for v in posibles if v}
    db = get_mongo_db()
    if db is None:
        flash(
            "No se pudo acceder a la base de datos. Contacte con el administrador.",
            "error",
        )
        return render_template(
            "error.html",
            mensaje="No se pudo conectar a la base de datos. Contacte con el administrador.",
        )
    try:
        tablas = []
        catalogos = []
        # Unificar criterio: buscar por todos los campos posibles
        query = {"$or": []}
        for val in posibles:
            query["$or"].extend(
                [
                    {"created_by": val},
                    {"owner": val},
                    {"owner_name": val},
                    {"email": val},
                    {"username": val},
                    {"name": val},
                ]
            )
        try:
            tablas = list(g.spreadsheets_collection.find(query).sort("created_at", -1))
        except Exception as e:
            print(f"[ERROR] Consulta a spreadsheets fall√≥: {e}")
            tablas = []
        try:
            catalogos = list(g.catalogs_collection.find(query).sort("created_at", -1))
        except Exception as e:
            print(f"[ERROR] Consulta a catalogs fall√≥: {e}")
            catalogos = []
        print(
            f"[DASHBOARD_USER] Tablas encontradas: {len(tablas)} | Cat√°logos encontrados: {len(catalogos)}"
        )

        # Refuerzo: normalizar y serializar campos
        def safe_str(val):
            if isinstance(val, datetime):
                return val.strftime("%Y-%m-%d %H:%M:%S")
            return str(val) if val is not None else ""

        current_app.logger.info(
            f"[DEBUG_TABLAS] Procesando {len(tablas)} tablas encontradas"
        )
        for t in tablas:
            current_app.logger.info(
                f"[DEBUG_TABLA_INDIVIDUAL] Procesando tabla: {t.get('name', 'Sin nombre')}"
            )
            t["tipo"] = "spreadsheet"
            # Unificaci√≥n: sincronizar 'rows' y 'data'
            if "rows" in t and t["rows"] is not None:
                t["data"] = t["rows"]
            elif "data" in t and t["data"] is not None:
                t["rows"] = t["data"]
            else:
                t["data"] = []
                t["rows"] = []
            t["row_count"] = len(t["rows"])
            t["_id"] = safe_str(t.get("_id"))
            t["created_at"] = safe_str(t.get("created_at"))
            t["owner"] = (
                t.get("owner")
                or t.get("created_by")
                or t.get("owner_name")
                or t.get("email")
                or t.get("username")
                or t.get("name")
                or ""
            )
            t["name"] = t.get("name", "")
            # üñºÔ∏è Miniatura: priorizar miniatura personalizada, luego buscar autom√°ticamente
            current_app.logger.info(
                f"[DEBUG_MINIATURA] Procesando tabla: {t.get('name', 'Sin nombre')}"
            )

            # 1. Verificar si tiene miniatura personalizada configurada
            if t.get("miniatura") and t["miniatura"].strip():
                current_app.logger.info(
                    f"[MINIATURA_CUSTOM] Usando miniatura personalizada: {t['miniatura']}"
                )
                # Ya est√° configurada, no hacer nada m√°s
                continue

            # 2. Si no tiene miniatura personalizada, buscar autom√°ticamente
            t["miniatura"] = ""

            # Buscar en todas las filas hasta encontrar una imagen
            for row in t.get("data", []):
                if not isinstance(row, dict):
                    continue

                imagen_encontrada = None

                # 1. Verificar campo "Imagen" (URLs externas como Unsplash)
                if (
                    row.get("Imagen")
                    and isinstance(row["Imagen"], str)
                    and row["Imagen"].startswith("http")
                ):
                    imagen_encontrada = row["Imagen"]

                # 2. Si no hay imagen externa, buscar en campos de im√°genes locales
                if not imagen_encontrada:
                    # Verificar campos de im√°genes locales
                    for campo in ["imagenes", "images", "imagen_data"]:
                        if campo in row and row[campo]:
                            imgs = (
                                row[campo]
                                if isinstance(row[campo], list)
                                else [row[campo]]
                            )
                            for img in imgs:
                                if (
                                    img
                                    and img != "N/A"
                                    and isinstance(img, str)
                                    and not img.startswith("http")
                                ):
                                    # Verificar si S3 est√° habilitado
                                    use_s3 = (
                                        os.environ.get("USE_S3", "false").lower()
                                        == "true"
                                    )
                                    if use_s3:
                                        from app.utils.s3_utils import get_s3_url

                                        s3_url = get_s3_url(img)
                                        if s3_url:
                                            imagen_encontrada = s3_url
                                        else:
                                            imagen_encontrada = url_for(
                                                "static", filename=f"uploads/{img}"
                                            )
                                    else:
                                        imagen_encontrada = url_for(
                                            "static", filename=f"uploads/{img}"
                                        )
                                    break
                        if imagen_encontrada:
                            break

                # Si encontramos una imagen, salir del bucle
                if imagen_encontrada:
                    t["miniatura"] = imagen_encontrada
                    current_app.logger.info(
                        f"[MINIATURA] Encontrada para tabla {t.get('name', 'Sin nombre')}: {imagen_encontrada}"
                    )
                    break

            # Si no se encontr√≥ ninguna imagen, dejar vac√≠o (se mostrar√° "‚Äî")
            if not t["miniatura"]:
                current_app.logger.info(
                    f"[MINIATURA] No se encontr√≥ imagen para tabla {t.get('name', 'Sin nombre')}"
                )
        for c in catalogos:
            c["tipo"] = "catalog"
            # Unificaci√≥n: sincronizar 'rows' y 'data'
            if "rows" in c and c["rows"] is not None:
                c["data"] = c["rows"]
            elif "data" in c and c["data"] is not None:
                c["rows"] = c["data"]
            else:
                c["data"] = []
                c["rows"] = []
            c["row_count"] = len(c["rows"])
            c["_id"] = safe_str(c.get("_id"))
            c["created_at"] = safe_str(c.get("created_at"))
            c["owner"] = (
                c.get("owner")
                or c.get("created_by")
                or c.get("owner_name")
                or c.get("email")
                or c.get("username")
                or c.get("name")
                or ""
            )
            c["name"] = c.get("name", "")
            # Miniatura: primera imagen de la primera fila
            c["miniatura"] = ""
            if c["data"] and isinstance(c["data"][0], dict):
                row = c["data"][0]
                imagenes = []
                if "imagenes" in row and row["imagenes"]:
                    imagenes = (
                        row["imagenes"]
                        if isinstance(row["imagenes"], list)
                        else [row["imagenes"]]
                    )
                elif "images" in row and row["images"]:
                    imagenes = (
                        row["images"]
                        if isinstance(row["images"], list)
                        else [row["images"]]
                    )
                elif "imagen" in row and row["imagen"]:
                    imagenes = (
                        row["imagen"]
                        if isinstance(row["imagen"], list)
                        else [row["imagen"]]
                    )
                if imagenes:
                    img = imagenes[0]
                    if isinstance(img, str) and img.startswith("http"):
                        c["miniatura"] = img
                    else:
                        # Verificar si S3 est√° habilitado
                        use_s3 = os.environ.get("USE_S3", "false").lower() == "true"
                        if use_s3:
                            from app.utils.s3_utils import get_s3_url

                            s3_url = get_s3_url(img)
                            if s3_url:
                                c["miniatura"] = s3_url
                            else:
                                c["miniatura"] = url_for(
                                    "static", filename=f"uploads/{img}"
                                )
                        else:
                            # Usar URL local directamente
                            c["miniatura"] = url_for(
                                "static", filename=f"uploads/{img}"
                            )
        registros = tablas + catalogos
        current_app.logger.info(
            f"[DEBUG_REGISTROS] Total registros a enviar al template: {len(registros)}"
        )
        for i, reg in enumerate(registros):
            current_app.logger.info(
                f"[DEBUG_REG_{i}] Nombre: {reg.get('name')}, Tipo: {reg.get('tipo')}, Miniatura: {reg.get('miniatura', 'NO_SET')}"
            )

        if not registros:
            flash("No tienes cat√°logos ni tablas asociados a tu usuario.", "info")
            return render_template("dashboard_unificado.html", registros=[])
        return render_template("dashboard_unificado.html", registros=registros)
    except Exception as e:
        print(f"[ERROR][DASHBOARD_USER] {e}")
        flash("Error al cargar tus cat√°logos/tablas.", "error")
        return render_template("dashboard_unificado.html", registros=[])


@main_bp.route("/editar/<id>", methods=["GET", "POST"])
def editar(id):
    # Eliminar verificaciones de sesi√≥n y permisos
    # if "username" not in session:
    #     return redirect(url_for("auth.login"))
    if "selected_table" not in session:
        # Si no hay tabla seleccionada, usar el ID proporcionado para buscar la tabla
        try:
            table_info = g.spreadsheets_collection.find_one(  # type: ignore
                {"_id": ObjectId(id)}
            )
            if table_info:
                selected_table = table_info.get("filename")
                session["selected_table"] = selected_table
            else:
                flash("Tabla no encontrada.", "error")
                return redirect(url_for("main.tables"))
        except Exception as e:
            logger.error(f"Error al buscar tabla por ID: {str(e)}")
            flash("Error al buscar la tabla.", "error")
            return redirect(url_for("main.tables"))
    else:
        selected_table = session["selected_table"]
        table_info = g.spreadsheets_collection.find_one(  # type: ignore
            {"filename": selected_table}
        )
        if not table_info:
            flash("Tabla no encontrada.", "error")
            return redirect(url_for("main.tables"))
    # Control de acceso: solo admin o due√±o puede editar
    if session.get("role") != "admin" and table_info.get("owner") != session.get(
        "username"
    ):
        flash("No tiene permisos para editar esta tabla", "error")
        return redirect(url_for("main.tables"))

    if request.method == "POST":
        nuevo_nombre = request.form.get("nombre", "").strip()
        nuevos_headers = [
            h.strip() for h in request.form.get("headers", "").split(",") if h.strip()
        ]
        update = {}
        if nuevo_nombre:
            update["name"] = nuevo_nombre
        if nuevos_headers:
            update["headers"] = nuevos_headers
        if update:
            g.spreadsheets_collection.update_one(
                {"filename": selected_table}, {"$set": update}
            )
            flash("Tabla actualizada correctamente.", "success")
        else:
            flash("No se detectaron cambios.", "info")
        return redirect(url_for("main.ver_tabla", table_id=id))

    # GET: mostrar formulario de edici√≥n
    return render_template("editar_tabla.html", table=table_info)


@main_bp.route("/ver_tabla/<table_id>")
@login_required
def ver_tabla(table_id):
    try:
        table = g.spreadsheets_collection.find_one({"_id": ObjectId(table_id)})

        # Asegurarse de que el propietario est√© disponible
        if "owner" not in table and "owner_name" in table:
            table["owner"] = table["owner_name"]
        elif "owner" not in table and "created_by" in table:
            table["owner"] = table["created_by"]
        elif "owner" not in table:
            table["owner"] = "Usuario desconocido"

        current_app.logger.info(
            f"[DEBUG][VISIONADO] Propietario de la tabla: {table.get('owner')}"
        )
        current_app.logger.info(
            f"[DEBUG][VISIONADO] Campos de la tabla: {list(table.keys())}"
        )

        # Asegurarse de que los datos est√©n disponibles
        if "data" not in table and "rows" in table:
            table["data"] = table["rows"]
            current_app.logger.info(
                f"[DEBUG][VISIONADO] Usando 'rows' como 'data', filas: {len(table['data'])}"
            )
        elif "data" not in table:
            table["data"] = []
            current_app.logger.info(
                "[DEBUG][VISIONADO] No se encontraron datos en la tabla"
            )
        if not table:
            flash("Tabla no encontrada.", "error")
            return redirect(url_for("main.dashboard_user"))

        # Log de sesi√≥n y permisos
        current_app.logger.info(f"[DEBUG][VISIONADO] Sesi√≥n: {dict(session)}")
        current_app.logger.info(
            f"[DEBUG][VISIONADO] table.owner: {table.get('owner')}, session.username: {session.get('username')}, session.role: {session.get('role')}"
        )

        # Normalizar el campo owner si est√° vac√≠o o es usuario_predeterminado
        username = session.get("username")
        role = session.get("role", "user")

        if (
            "owner" not in table
            or not table["owner"]
            or table["owner"] == "usuario_predeterminado"
        ):
            # Buscar el propietario en otros campos
            if "created_by" in table and table["created_by"]:
                table["owner"] = table["created_by"]
                current_app.logger.info(
                    f"[DEBUG][VISIONADO] Usando created_by como owner: {table['owner']}"
                )
            elif "username" in table and table["username"]:
                table["owner"] = table["username"]
                current_app.logger.info(
                    f"[DEBUG][VISIONADO] Usando username como owner: {table['owner']}"
                )
            else:
                # Si no hay informaci√≥n de propietario, asignar al usuario actual
                g.spreadsheets_collection.update_one(
                    {"_id": ObjectId(table_id)}, {"$set": {"owner": username}}
                )
                table["owner"] = username
                current_app.logger.info(
                    f"[DEBUG][VISIONADO] Actualizado propietario de tabla {table_id} a {username}"
                )

        # Verificar permisos: solo el propietario o admin puede ver la tabla
        if role != "admin" and table.get("owner") != username:
            mensaje = (
                f"No tiene permisos para ver esta tabla. "
                f"(owner={table.get('owner')}, username={username}, role={role})"
            )
            current_app.logger.warning(f"[DEBUG][VISIONADO] {mensaje}")
            flash(mensaje, "warning")
            return redirect(url_for("main.tables"))

        # Si llegamos aqu√≠, el usuario tiene permisos para ver la tabla
        # Procesar las im√°genes en cada fila usando funci√≥n unificada
        for i, row in enumerate(table.get("data", [])):
            if not isinstance(row, dict):
                current_app.logger.warning(
                    f"[VISIONADO] Fila {i} ignorada por no ser un dict: {row}"
                )
                continue

            current_app.logger.info(f"[DEBUG][VISIONADO] Procesando fila {i}: {row}")

            # DEBUG: Verificar campos Documentaci√≥n_0, Documentaci√≥n_1, etc.
            for key, value in row.items():
                if key.startswith("Documentaci√≥n_"):
                    current_app.logger.info(f"[DEBUG][VISIONADO] Campo {key}: {value}")

            # Limpiar cache de S3 al cargar la p√°gina para evitar im√°genes fantasma
            from app.utils.image_utils import clear_s3_cache

            clear_s3_cache()  # Limpiar todo el cache

            # Usar funci√≥n unificada para obtener URLs de im√°genes
            image_data = get_images_for_template(row)
            row.update(image_data)  # A√±ade imagen_urls, num_imagenes, tiene_imagenes

            current_app.logger.info(
                f"[DEBUG][VISIONADO] URLs de im√°genes para fila {i}: {row.get('imagen_urls', [])}"
            )
            current_app.logger.info(
                f"[DEBUG][VISIONADO] Total de im√°genes en fila {i}: {len(row.get('imagen_urls', []))}"
            )

            # Procesar campos de Documentaci√≥n - crear URLs correctas para documentos
            from app.utils.s3_utils import get_s3_url

            for key, value in row.items():
                if key.startswith("Documentaci√≥n_") and value:
                    current_app.logger.info(
                        f"[DEBUG][VISIONADO] Procesando documento {key}: {value}"
                    )

                    # Verificar si ya es una URL completa
                    if isinstance(value, str) and (
                        value.startswith("/admin/s3/")
                        or value.startswith("/static/uploads/")
                        or value.startswith("/imagenes_subidas/")
                        or value.startswith("http")
                    ):
                        # Ya es una URL completa, usar directamente
                        current_app.logger.info(
                            f"[DEBUG][VISIONADO] URL completa detectada para {key}: {value}"
                        )
                        continue

                    # Si es solo un nombre de archivo, intentar obtener la URL de S3
                    # primero
                    if isinstance(value, str) and len(value) > 5:
                        s3_url = get_s3_url(value)
                        if s3_url:
                            # Convertir URL S3 a proxy local
                            filename = value.split("/")[-1] if "/" in value else value
                            proxy_url = f"/admin/s3/{filename}"
                            row[key] = proxy_url
                            current_app.logger.info(
                                f"[DEBUG][VISIONADO] Documento S3 encontrado: {value} -> {proxy_url}"
                            )
                        else:
                            # Si no est√° en S3, usar la URL local
                            # Verificar si ya contiene una ruta para evitar
                            # concatenaci√≥n incorrecta
                            if (
                                value.startswith("/admin/s3/")
                                or value.startswith("/static/uploads/")
                                or value.startswith("/imagenes_subidas/")
                            ):
                                local_url = value
                            else:
                                local_url = f"/static/uploads/{value}"
                            row[key] = local_url
                            current_app.logger.info(
                                f"[DEBUG][VISIONADO] Usando URL local para documento: {value} -> {local_url}"
                            )
                    elif isinstance(value, list):
                        # Si es una lista de documentos, procesar cada uno
                        processed_docs = []
                        for doc in value:
                            if isinstance(doc, str) and len(doc) > 5:
                                s3_url = get_s3_url(doc)
                                if s3_url:
                                    filename = doc.split("/")[-1] if "/" in doc else doc
                                    proxy_url = f"/admin/s3/{filename}"
                                    processed_docs.append(proxy_url)
                                    current_app.logger.info(
                                        f"[DEBUG][VISIONADO] Documento S3 en lista: {doc} -> {proxy_url}"
                                    )
                                else:
                                    # Verificar si ya contiene una ruta para evitar
                                    # concatenaci√≥n incorrecta
                                    if (
                                        doc.startswith("/admin/s3/")
                                        or doc.startswith("/static/uploads/")
                                        or doc.startswith("/imagenes_subidas/")
                                    ):
                                        local_url = doc
                                    else:
                                        local_url = f"/static/uploads/{doc}"
                                    processed_docs.append(local_url)
                                    current_app.logger.info(
                                        f"[DEBUG][VISIONADO] Documento local en lista: {doc} -> {local_url}"
                                    )
                            else:
                                processed_docs.append(doc)
                        row[key] = processed_docs

            # Procesar campo Multimedia - crear URL correcta
            if "Multimedia" in row and row["Multimedia"]:
                multimedia_value = row["Multimedia"]
                current_app.logger.info(
                    f"[DEBUG][VISIONADO] Procesando multimedia: {multimedia_value}"
                )

                # Verificar si ya es una URL completa
                if isinstance(multimedia_value, str) and (
                    multimedia_value.startswith("/admin/s3/")
                    or multimedia_value.startswith("/static/uploads/")
                    or multimedia_value.startswith("/imagenes_subidas/")
                    or multimedia_value.startswith("http")
                ):
                    # Ya es una URL completa, usar directamente
                    current_app.logger.info(
                        f"[DEBUG][VISIONADO] URL multimedia completa detectada: {multimedia_value}"
                    )
                elif isinstance(multimedia_value, str) and len(multimedia_value) > 5:
                    # Si es solo un nombre de archivo, intentar obtener la URL de S3
                    # primero
                    s3_url = get_s3_url(multimedia_value)
                    if s3_url:
                        # Convertir URL S3 a proxy local
                        filename = (
                            multimedia_value.split("/")[-1]
                            if "/" in multimedia_value
                            else multimedia_value
                        )
                        proxy_url = f"/admin/s3/{filename}"
                        row["Multimedia"] = proxy_url
                        current_app.logger.info(
                            f"[DEBUG][VISIONADO] Multimedia S3 encontrado: {multimedia_value} -> {proxy_url}"
                        )
                    else:
                        # Si no est√° en S3, usar la URL local directamente
                        # No usar proxy S3 para archivos locales
                        # Verificar si ya contiene una ruta para evitar concatenaci√≥n
                        # incorrecta
                        if (
                            multimedia_value.startswith("/admin/s3/")
                            or multimedia_value.startswith("/static/uploads/")
                            or multimedia_value.startswith("/imagenes_subidas/")
                        ):
                            local_url = multimedia_value
                        else:
                            local_url = f"/static/uploads/{multimedia_value}"
                        row["Multimedia"] = local_url
                        current_app.logger.info(
                            f"[DEBUG][VISIONADO] Usando URL local para multimedia: {multimedia_value} -> {local_url}"
                        )

        return render_template("ver_tabla.html", table=table)
    except BuildError as e:
        logger.error(f"BuildError en ver_tabla: {str(e)}", exc_info=True)
        flash("Error interno: ruta no encontrada o mal configurada.", "danger")
        return render_template(
            "error.html", error="Error interno: ruta no encontrada o mal configurada."
        )
    except Exception as e:
        logger.error(f"Error inesperado en ver_tabla: {str(e)}", exc_info=True)
        flash("Error interno inesperado.", "danger")
        return redirect(url_for("main.dashboard_user"))


@main_bp.route("/select_table/<table_id>")
def select_table(table_id):
    # Verificar que el usuario ha iniciado sesi√≥n
    if not session.get("username"):
        flash("Debe iniciar sesi√≥n para acceder a las tablas", "warning")
        return redirect(url_for("auth.login"))

    # Obtener informaci√≥n del usuario actual
    username = session.get("username")
    role = session.get("role", "user")

    try:
        # Obtener la tabla
        table = g.spreadsheets_collection.find_one({"_id": ObjectId(table_id)})
        if not table:
            flash("Tabla no encontrada.", "error")
            return redirect(url_for("main.tables"))

        # Verificar permisos: administradores pueden ver todas las tablas,
        # usuarios solo las suyas
        if role != "admin":
            # Normalizar el campo owner si est√° vac√≠o o es usuario_predeterminado
            if (
                "owner" not in table
                or not table["owner"]
                or table["owner"] == "usuario_predeterminado"
            ):
                if "created_by" in table and table["created_by"]:
                    table["owner"] = table["created_by"]
                elif "username" in table and table["username"]:
                    table["owner"] = table["username"]
                else:
                    # Si no hay informaci√≥n de propietario, asignar al usuario actual
                    g.spreadsheets_collection.update_one(
                        {"_id": ObjectId(table_id)}, {"$set": {"owner": username}}
                    )
                    table["owner"] = username
                    logger.info(
                        f"Actualizado propietario de tabla {table_id} a {username}"
                    )

            # Verificar si el usuario actual es el propietario
            if table["owner"] != username:
                logger.warning(
                    f"Usuario {username} intent√≥ acceder a tabla {table_id} propiedad de {table['owner']}"
                )
                flash("No tiene permisos para ver esta tabla.", "warning")
                return redirect(url_for("main.tables"))

        # Guardar informaci√≥n de la tabla en la sesi√≥n
        session["selected_table"] = table.get("filename", "")
        session["selected_table_id"] = str(table["_id"])
        session["selected_table_name"] = table.get("name", "Sin nombre")

        logger.info(
            f"Usuario {username} seleccion√≥ tabla {table.get('name')} (ID: {table_id})"
        )
        return redirect(url_for("main.ver_tabla", table_id=table_id))

    except Exception as e:
        logger.error(f"Error al seleccionar tabla {table_id}: {str(e)}", exc_info=True)
        flash(f"Error al acceder a la tabla: {str(e)}", "error")
        return redirect(url_for("main.tables"))


@main_bp.route("/perfil")
def perfil():
    # Verificar si el usuario est√° autenticado
    if "user_id" not in session:
        flash("Debe iniciar sesi√≥n para ver su perfil", "warning")
        return redirect(url_for("auth.login"))

    try:
        # Obtener la colecci√≥n de usuarios
        users_collection = getattr(g, "users_collection", None)
        if users_collection is None:
            # Importar mongo desde app.extensions
            from app.extensions import mongo

            users_collection = mongo.db.users  # type: ignore

        # Obtener datos del usuario actual
        user_data = users_collection.find_one({"_id": ObjectId(session["user_id"])})
        if not user_data:
            flash("Usuario no encontrado", "error")
            return redirect(url_for("main.dashboard_user"))

        # Crear objeto User para tener acceso a foto_perfil_url
        from app.models.user import User

        user = User(user_data)

        # Asegurar que existe la imagen de perfil predeterminada
        if current_app.static_folder is None:
            flash("Error de configuraci√≥n del servidor", "error")
            return redirect(url_for("main.dashboard_user"))

        default_profile_path = os.path.join(
            current_app.static_folder, "default_profile.png"
        )
        if not os.path.exists(default_profile_path):
            try:
                # Importar la funci√≥n para crear la imagen predeterminada
                from app.crear_imagen_perfil_default import crear_imagen_perfil_default

                crear_imagen_perfil_default()
                logger.info("Imagen de perfil predeterminada creada correctamente")
            except Exception as e:
                logger.error(
                    f"Error al crear imagen predeterminada: {str(e)}", exc_info=True
                )

        return render_template("perfil.html", user=user)
    except Exception as e:
        logger.error(f"Error al cargar perfil: {str(e)}", exc_info=True)
        flash("Error al cargar los datos del perfil", "error")
        return redirect(url_for("main.dashboard_user"))


@main_bp.route("/editar_perfil", methods=["GET", "POST"])
def editar_perfil():
    if "user_id" not in session:
        flash("Debe iniciar sesi√≥n para editar su perfil", "warning")
        return redirect(url_for("auth.login"))

    # Obtener la colecci√≥n de usuarios
    users_collection = getattr(g, "users_collection", None)
    if users_collection is None:
        # Importar mongo desde app.extensions
        from app.extensions import mongo

        if mongo.db is None:
            flash("Error de conexi√≥n a la base de datos.", "error")
            return redirect(url_for("main.dashboard_user"))

        users_collection = mongo.db.users

    # Obtener datos del usuario actual
    try:
        user_data = users_collection.find_one({"_id": ObjectId(session["user_id"])})
        if not user_data:
            flash("Usuario no encontrado", "error")
            return redirect(url_for("main.dashboard_user"))

        # Crear objeto User para tener acceso a foto_perfil_url
        from app.models.user import User

        user = User(user_data)
    except Exception as e:
        logger.error(f"Error al obtener datos del usuario: {str(e)}", exc_info=True)
        flash("Error al cargar los datos del usuario", "error")
        return redirect(url_for("main.dashboard_user"))

    if request.method == "POST":
        # Obtener datos del formulario
        nombre = request.form.get("nombre")
        email = request.form.get("email")

        # Datos para cambio de contrase√±a
        password_actual = request.form.get("password_actual")
        password_nuevo = request.form.get("password_nuevo")
        password_confirmacion = request.form.get("password_confirmacion")

        # Actualizar en la base de datos
        update_data = {}
        if nombre:
            update_data["nombre"] = nombre
        if email:
            update_data["email"] = email

        # Procesar cambio de contrase√±a si se proporcionaron los campos necesarios
        if password_actual and password_nuevo and password_confirmacion:
            # Verificar que la contrase√±a actual sea correcta
            from werkzeug.security import check_password_hash, generate_password_hash

            if not check_password_hash(user.password_hash, password_actual):
                flash("La contrase√±a actual es incorrecta.", "error")
                return render_template("editar_perfil.html", user=user)

            # Verificar que las nuevas contrase√±as coincidan
            if password_nuevo != password_confirmacion:
                flash("Las nuevas contrase√±as no coinciden.", "error")
                return render_template("editar_perfil.html", user=user)

            # Verificar que la nueva contrase√±a tenga al menos 8 caracteres
            if len(password_nuevo) < 8:
                flash("La nueva contrase√±a debe tener al menos 8 caracteres.", "error")
                return render_template("editar_perfil.html", user=user)

            # Actualizar la contrase√±a
            update_data["password"] = generate_password_hash(password_nuevo)

        # Asegurarse de que existe la carpeta de uploads
        if current_app.static_folder is None:
            flash("Error de configuraci√≥n del servidor", "error")
            return redirect(url_for("main.dashboard_user"))

        uploads_folder = os.path.join(current_app.static_folder, "uploads")

        # Crear la carpeta si no existe
        os.makedirs(uploads_folder, exist_ok=True)
        # Asegurar que existe la imagen de perfil predeterminada
        default_profile_path = os.path.join(
            current_app.static_folder, "default_profile.png"
        )
        if not os.path.exists(default_profile_path):
            try:
                # Importar la funci√≥n para crear la imagen predeterminada
                from app.crear_imagen_perfil_default import crear_imagen_perfil_default

                crear_imagen_perfil_default()
                logger.info("Imagen de perfil predeterminada creada correctamente")
            except Exception as e:
                logger.error(
                    f"Error al crear imagen predeterminada: {str(e)}", exc_info=True
                )

        # Procesar la imagen de perfil si se ha subido una nueva
        if "foto" in request.files:
            profile_image = request.files["foto"]
            if profile_image and profile_image.filename != "":
                try:
                    # Guardar la imagen
                    filename = secure_filename(
                        f"{uuid.uuid4().hex}_{profile_image.filename}"
                    )
                    filepath = os.path.join(uploads_folder, filename)
                    profile_image.save(filepath)

                    # Subir a S3 si est√° habilitado
                    use_s3 = os.environ.get("USE_S3", "false").lower() == "true"
                    if use_s3:
                        try:
                            from app.utils.s3_utils import upload_file_to_s3

                            logger.info(f"Subiendo foto de perfil a S3: {filename}")
                            result = upload_file_to_s3(filepath, filename)
                            if result["success"]:
                                logger.info(
                                    f"Foto de perfil subida a S3: {result['url']}"
                                )
                                # Eliminar el archivo local despu√©s de subirlo a S3
                                os.remove(filepath)
                                logger.info(
                                    f"Foto de perfil eliminada del servidor local despu√©s de subirla a S3: {filename}"
                                )
                            else:
                                logger.error(
                                    f"Error al subir foto de perfil a S3: {result.get('error')}"
                                )
                        except Exception as e:
                            logger.error(
                                f"Error al procesar subida de foto de perfil a S3: {str(e)}",
                                exc_info=True,
                            )
                    # Actualizar el campo foto_perfil en el usuario
                    update_data["foto_perfil"] = filename
                    logger.info(f"Imagen de perfil guardada: {filename}")
                except Exception as e:
                    logger.error(
                        f"Error al guardar imagen de perfil: {str(e)}", exc_info=True
                    )
                    flash(f"Error al guardar la imagen de perfil: {str(e)}", "error")

        # Actualizar el usuario en la base de datos
        from app.extensions import mongo

        if mongo.db is None:
            flash("Error de conexi√≥n a la base de datos.", "error")
            return redirect(url_for("main.editar_perfil"))

        mongo.db.users.update_one(
            {"_id": ObjectId(session["user_id"])}, {"$set": update_data}
        )

        # Mostrar mensaje espec√≠fico si se cambi√≥ la contrase√±a
        if "password" in update_data:
            flash("Perfil y contrase√±a actualizados correctamente.", "success")
            # Actualizar la sesi√≥n para reflejar los cambios
            if "email" in update_data:
                session["email"] = update_data["email"]
            if "nombre" in update_data and update_data["nombre"]:
                session["username"] = update_data["nombre"]
            # Redirigir al login para que el usuario inicie sesi√≥n con la nueva
            # contrase√±a
            session.clear()
            flash("Por favor, inicie sesi√≥n con su nueva contrase√±a.", "info")
            return redirect(url_for("auth.login"))
        else:
            flash("Perfil actualizado correctamente.", "success")
            # Actualizar la sesi√≥n para reflejar los cambios
            if "email" in update_data:
                session["email"] = update_data["email"]
            if "nombre" in update_data and update_data["nombre"]:
                session["username"] = update_data["nombre"]
            return redirect(url_for("main.perfil"))

    # Para peticiones GET, mostrar el formulario con los datos actuales
    return render_template("editar_perfil.html", user=user)


@main_bp.route("/editar_fila/<tabla_id>/<int:fila_index>", methods=["GET", "POST"])
@login_required
def editar_fila(tabla_id, fila_index):
    from datetime import datetime

    print(f"üî•üî•üî• EDITAR_FILA EJECUTADO: tabla_id={tabla_id}, fila_index={fila_index}")
    current_app.logger.error(
        f"üî•üî•üî• EDITAR_FILA EJECUTADO: tabla_id={tabla_id}, fila_index={fila_index}"
    )
    current_app.logger.info(
        f"[DEBUG] Valores recibidos: tabla_id={tabla_id}, fila_index={fila_index}"
    )

    # üî• OBTENER INFO FRESCA USANDO LA MISMA L√ìGICA QUE VER_TABLA üî•
    current_app.logger.info(
        f"[DEBUG_EDIT] Recargando datos frescos desde MongoDB para tabla {tabla_id}"
    )

    # USAR EXACTAMENTE LA MISMA L√ìGICA QUE VER_TABLA
    table_info = g.spreadsheets_collection.find_one({"_id": ObjectId(tabla_id)})
    if not table_info:
        flash("Tabla no encontrada.", "error")
        return redirect(url_for("main.tables"))

    # Asegurarse de que el propietario est√© disponible (igual que ver_tabla)
    if "owner" not in table_info and "owner_name" in table_info:
        table_info["owner"] = table_info["owner_name"]
    elif "owner" not in table_info and "created_by" in table_info:
        table_info["owner"] = table_info["created_by"]
    elif "owner" not in table_info:
        table_info["owner"] = "Usuario desconocido"

    # Asegurarse de que los datos est√©n disponibles (igual que ver_tabla)
    if "data" not in table_info:
        if "rows" in table_info:
            table_info["data"] = table_info["rows"]
            current_app.logger.info(
                f"[DEBUG_EDIT] Usando 'rows' como 'data', filas: {len(table_info['data'])}"
            )
        else:
            table_info["data"] = []
            current_app.logger.info("[DEBUG_EDIT] No se encontraron datos en la tabla")

    # üö® NO SOBRESCRIBIR DATA CON ROWS - data tiene las im√°genes actualizadas
    current_app.logger.info(
        f"[DEBUG_EDIT] ANTES sincronizaci√≥n - data tiene {len(table_info.get('data', []))} filas, rows tiene {len(table_info.get('rows', []))} filas"
    )

    # Solo sincronizar si data est√° vac√≠o pero rows tiene datos
    if (
        not table_info.get("data") or len(table_info.get("data", [])) == 0
    ) and table_info.get("rows"):
        current_app.logger.info("[DEBUG_EDIT] Copiando rows ‚Üí data (data estaba vac√≠o)")
        table_info["data"] = table_info["rows"]
    elif table_info.get("data") and (
        not table_info.get("rows") or len(table_info.get("rows", [])) == 0
    ):
        current_app.logger.info("[DEBUG_EDIT] Copiando data ‚Üí rows (rows estaba vac√≠o)")
        table_info["rows"] = table_info["data"]
    else:
        current_app.logger.info("[DEBUG_EDIT] PRESERVANDO data original con im√°genes")

    table_info["row_count"] = len(table_info.get("data", []))

    # Verificar permisos: solo el propietario o admin puede editar filas
    username = session.get("username")
    role = session.get("role", "user")

    # Obtener el propietario de la tabla (puede estar en diferentes campos)
    owner = (
        table_info.get("owner")
        or table_info.get("created_by")
        or table_info.get("owner_name")
    )

    current_app.logger.info(
        f"[DEBUG] Verificando permisos: role={role}, username={username}, owner={owner}"
    )

    if role != "admin" and owner != username:
        current_app.logger.warning(
            f"[DEBUG] Permiso denegado: {username} intentando editar tabla de {owner}"
        )
        flash("No tienes permisos para editar esta fila.", "warning")
        return redirect(url_for("main.ver_tabla", table_id=tabla_id))

    # Obtener la fila espec√≠fica del array de datos
    data = table_info.get("data", {})

    # Verificar si data es un diccionario o una lista
    if isinstance(data, dict):
        # Si es un diccionario, usar la clave string
        data_key = str(fila_index)
        if data_key not in data:
            current_app.logger.error(
                f"[DEBUG] Fila no encontrada: clave {data_key} no existe en data"
            )
            flash("Fila no encontrada.", "error")
            return redirect(url_for("main.ver_tabla", table_id=tabla_id))
        fila = data[data_key]
        data_length = len(data)
    else:
        # Si es una lista, usar el √≠ndice num√©rico
        data_length = len(data)
        if fila_index >= data_length:
            current_app.logger.error(
                f"[DEBUG] Fila no encontrada: √≠ndice {fila_index} >= longitud {data_length}"
            )
            flash("Fila no encontrada.", "error")
            return redirect(url_for("main.ver_tabla", table_id=tabla_id))
        fila = data[fila_index]

    current_app.logger.info(
        f"[DEBUG] fila_index={fila_index}, data_length={data_length}, data_type={type(data)}"
    )
    current_app.logger.info(f"[DEBUG] table_info.keys(): {list(table_info.keys())}")
    current_app.logger.info(
        f"[DEBUG_EDIT] Datos completos de fila {fila_index}: {fila}"
    )

    # üî• DEBUG ADICIONAL: Verificar qu√© contiene cada campo de imagen üî•
    current_app.logger.info(
        f"[DEBUG_EDIT] fila.get('images'): {fila.get('images')} (tipo: {type(fila.get('images'))})"
    )
    current_app.logger.info(
        f"[DEBUG_EDIT] fila.get('imagenes'): {fila.get('imagenes')} (tipo: {type(fila.get('imagenes'))})"
    )
    current_app.logger.info(
        f"[DEBUG_EDIT] fila.get('imagen_data'): {fila.get('imagen_data')} (tipo: {type(fila.get('imagen_data'))})"
    )
    current_app.logger.info(
        f"[DEBUG_EDIT] Campos de imagen disponibles: imagenes={fila.get('imagenes')}, imagen_data={fila.get('imagen_data')}, images={fila.get('images')}"
    )

    # üî•üî•üî• USAR LA MISMA L√ìGICA QUE VER_TABLA CON CACHE üî•üî•üî•
    # Limpiar cache de S3 al cargar la p√°gina para evitar im√°genes fantasma
    from app.utils.image_utils import clear_s3_cache

    clear_s3_cache()  # Limpiar todo el cache

    # Optimizaci√≥n: Limpiar cache de S3 solo si es necesario
    current_app.logger.info(
        "[DEBUG_EDIT] Cache S3 limpiado para evitar im√°genes fantasma"
    )

    # Usar funci√≥n unificada para obtener URLs de im√°genes (filtra las que no existen)
    image_data = get_images_for_template(fila)
    fila.update(image_data)  # A√±ade imagen_urls, num_imagenes, tiene_imagenes

    # Para editar, necesitamos los nombres de archivo originales (no URLs)
    # pero solo los que realmente existen
    imagenes_existentes = []
    for img_url in image_data.get("imagen_urls", []):
        # Extraer nombre de archivo de la URL
        if "/static/uploads/" in img_url:
            filename = img_url.split("/static/uploads/")[-1]
            imagenes_existentes.append(filename)
        elif "/admin/s3/" in img_url:
            filename = img_url.split("/admin/s3/")[-1]
            imagenes_existentes.append(filename)

    # Actualizar campos para consistencia
    if imagenes_existentes:
        fila["images"] = imagenes_existentes
        # Eliminar campos duplicados para evitar confusi√≥n
        if "imagenes" in fila:
            del fila["imagenes"]
        if "imagen_data" in fila:
            del fila["imagen_data"]

    # üÜï RECOPILAR DE FILA.IMAGEN (CAMPO STRING) - PARA IM√ÅGENES UNSPLASH
    imagen_individual = fila.get("Imagen", "")
    if (
        imagen_individual
        and imagen_individual != "N/A"
        and isinstance(imagen_individual, str)
        and imagen_individual.startswith("http")
    ):
        current_app.logger.info(
            f"[DEBUG_EDIT] Imagen externa encontrada: {imagen_individual}"
        )
        # Las im√°genes externas (Unsplash) no se pueden eliminar, pero necesitamos mostrarlas
        # Para eso, creamos un campo especial para mostrar sin opci√≥n de eliminar

    # Preparar contexto para el template con im√°genes unificadas (solo las que existen)
    fila["_imagenes_unificadas"] = imagenes_existentes
    fila["_imagen_externa"] = (
        imagen_individual
        if imagen_individual
        and isinstance(imagen_individual, str)
        and imagen_individual.startswith("http")
        else None
    )
    current_app.logger.info(
        f"[DEBUG_EDIT] IM√ÅGENES EXISTENTES: {len(imagenes_existentes)} ‚Üí {imagenes_existentes}"
    )
    current_app.logger.error(
        f"[DEBUG_EDIT] TEMPLATE CONTEXT: Todos los keys de fila = {list(fila.keys())}"
    )
    headers = table_info.get("headers", [])

    if request.method == "POST":
        current_app.logger.info("[DEBUG_EDIT] Iniciando procesamiento POST")
        update_data = {}

        # Procesar todos los documentos primero (antes del bucle normal)
        # Buscar todos los campos de documentos con el patr√≥n Documentaci√≥n_file_INDEX
        current_app.logger.info(
            f"[DEBUG_EDIT] Procesando {len(request.files)} archivos"
        )
        current_app.logger.info(
            f"[DEBUG_EDIT] request.files.keys(): {list(request.files.keys())}"
        )
        try:
            for key, file in request.files.items():
                current_app.logger.info(f"[DEBUG_EDIT] Procesando archivo: {key}")
                # Detectar archivos de documentos con formato Documentaci√≥n_X_file_Y
                is_document_file = key.startswith("Documentaci√≥n_file_") or (
                    key.startswith("Documentaci√≥n_") and "_file_" in key
                )
                current_app.logger.info(
                    f"[DEBUG_EDIT] ¬øEs archivo de documento? {is_document_file} para {key}"
                )
                if is_document_file and file.filename:
                    current_app.logger.info(
                        f"[DEBUG_EDIT] Archivo de documentaci√≥n encontrado: {file.filename}"
                    )
                    # Extraer el √≠ndice del nombre del campo
                    if key.startswith("Documentaci√≥n_file_"):
                        # Formato: Documentaci√≥n_file_X
                        index_str = key.replace("Documentaci√≥n_file_", "")
                    else:
                        # Formato: Documentaci√≥n_X_file_Y
                        # Extraer el n√∫mero despu√©s de Documentaci√≥n_ y antes de _file_
                        parts = key.split("_")
                        if len(parts) >= 3 and parts[0] == "Documentaci√≥n":
                            index_str = parts[1]
                        else:
                            current_app.logger.error(
                                f"[DEBUG_EDIT] No se pudo extraer √≠ndice de {key}"
                            )
                            continue

                    try:
                        column_index = int(index_str)
                        current_app.logger.info(
                            f"[DEBUG_EDIT] √çndice de columna: {column_index}"
                        )

                        # Procesar archivo documento
                        filename = secure_filename(
                            f"{uuid.uuid4().hex}_{file.filename}"
                        )
                        upload_dir = get_upload_dir()
                        file_path = os.path.join(upload_dir, filename)
                        file.save(file_path)

                        # Subir a S3 si est√° habilitado
                        use_s3 = os.environ.get("USE_S3", "false").lower() == "true"
                        if use_s3:
                            try:
                                from app.utils.s3_utils import upload_file_to_s3

                                current_app.logger.info(
                                    f"Subiendo documento a S3: {filename}"
                                )
                                result = upload_file_to_s3(file_path, filename)
                                if result and result.get("success"):
                                    current_app.logger.info(
                                        f"Documento subido exitosamente a S3: {filename}"
                                    )
                                    # Eliminar archivo local despu√©s de subirlo a S3
                                    os.remove(file_path)
                                    campo_valor = filename
                                else:
                                    error_msg = (
                                        result.get("error", "Error desconocido")
                                        if result
                                        else "Resultado None"
                                    )
                                    current_app.logger.error(
                                        f"Error subiendo documento a S3: {error_msg}"
                                    )
                                    # Si falla S3, mantener archivo local
                                    campo_valor = filename
                            except Exception as e:
                                current_app.logger.error(
                                    f"Error subiendo documento a S3 {filename}: {e}"
                                )
                                # Si falla S3, mantener archivo local
                                campo_valor = filename
                        else:
                            campo_valor = filename

                        # Guardar en el campo espec√≠fico Documentaci√≥n_X
                        header_name = f"Documentaci√≥n_{column_index}"
                        update_data[f"data.{fila_index}.{header_name}"] = campo_valor
                        current_app.logger.info(
                            f"[DEBUG_EDIT] Documento guardado en {header_name}: {campo_valor}"
                        )

                    except ValueError:
                        current_app.logger.error(
                            f"√çndice inv√°lido en campo de documento: {key}"
                        )
        except Exception as e:
            current_app.logger.error(
                f"[DEBUG_EDIT] Error en procesamiento de archivos: {e}"
            )

        # Procesar URLs de documentos
        for key, value in request.form.items():
            if key.startswith("Documentaci√≥n_url_") and value.strip():
                # Extraer el √≠ndice del nombre del campo
                index_str = key.replace("Documentaci√≥n_url_", "")
                try:
                    column_index = int(index_str)

                    # Guardar en el campo espec√≠fico Documentaci√≥n_X
                    header_name = f"Documentaci√≥n_{column_index}"
                    update_data[f"data.{fila_index}.{header_name}"] = value.strip()
                    current_app.logger.info(
                        f"[DEBUG_EDIT] URL de documento guardada en {header_name}: {value.strip()}"
                    )

                except ValueError:
                    current_app.logger.error(
                        f"√çndice inv√°lido en campo de documento: {key}"
                    )

        for header in headers:
            if header == "Multimedia":
                # Manejar campo Multimedia
                multimedia_url = request.form.get(f"{header}_url", "").strip()
                multimedia_file = request.files.get(f"{header}_file")

                if multimedia_url:
                    campo_valor = multimedia_url
                elif multimedia_file and multimedia_file.filename:
                    # Procesar archivo multimedia
                    filename = secure_filename(
                        f"{uuid.uuid4().hex}_{multimedia_file.filename}"
                    )
                    upload_dir = get_upload_dir()
                    file_path = os.path.join(upload_dir, filename)
                    multimedia_file.save(file_path)
                    campo_valor = filename
                else:
                    # Mantener valor existente si no hay cambios
                    campo_valor = fila.get(header, "")

            elif header == "Documentaci√≥n":
                # Los documentos ya se procesaron arriba, saltar
                continue

            elif header == "Fecha":
                # Columna inteligente: mantener fecha existente o asignar nueva si est√°
                # vac√≠a
                fecha_valor = request.form.get(header, "").strip()
                if not fecha_valor:
                    # Si no se proporciona fecha, mantener la existente o asignar fecha
                    # actual
                    campo_valor = fila.get(header, "")
                    if not campo_valor:
                        campo_valor = datetime.now().strftime("%Y-%m-%d")
                else:
                    campo_valor = fecha_valor
            elif header != "N√∫mero" and header != "Imagenes":
                # Campo normal - PRESERVAR valores existentes si no se proporcionan nuevos
                # PERO NO sobrescribir campos de Documentaci√≥n que ya fueron procesados
                if header.startswith("Documentaci√≥n_"):
                    # Verificar si ya se proces√≥ este campo de Documentaci√≥n
                    document_key = f"data.{fila_index}.{header}"
                    if document_key in update_data:
                        # Ya se proces√≥, no sobrescribir
                        current_app.logger.info(
                            f"[DEBUG_EDIT] Campo {header} ya procesado, saltando preservaci√≥n"
                        )
                        continue

                form_value = request.form.get(header, "").strip()
                if form_value:
                    # Si hay un valor en el formulario, usarlo
                    campo_valor = form_value
                else:
                    # Si no hay valor en el formulario, preservar el valor existente
                    campo_valor = fila.get(header, "")
                    current_app.logger.info(
                        f"[DEBUG_EDIT] Preservando valor existente para {header}: {campo_valor}"
                    )

            # Usamos notaci√≥n de diccionario para manejar campos con espacios
            # En lugar de usar dot notation directamente en la clave
            data_key = f"data.{fila_index}"

            # Si el campo no existe en update_data, creamos un diccionario vac√≠o
            if data_key not in update_data:
                update_data[data_key] = {}

            # Agregar el campo al diccionario
            update_data[data_key][header] = campo_valor

        # üÜï PROCESAR ESPEC√çFICAMENTE EL CAMPO 'Imagen' (URL EXTERNA)
        imagen_url = request.form.get("Imagen", "").strip()
        data_key = f"data.{fila_index}"
        if data_key not in update_data:
            update_data[data_key] = {}
        update_data[data_key]["Imagen"] = imagen_url
        current_app.logger.info(
            f"[DEBUG_EDIT] Actualizando campo Imagen: '{imagen_url}'"
        )

        # Procesar im√°genes a eliminar
        imagenes_a_eliminar = request.form.get("imagenes_a_eliminar", "")
        if imagenes_a_eliminar:
            try:
                import json

                imagenes_a_eliminar = json.loads(imagenes_a_eliminar)
                logger.info(f"Im√°genes a eliminar: {imagenes_a_eliminar}")

                # Asegurarse de que fila tenga el campo 'imagenes'
                if "imagenes" not in fila:
                    fila["imagenes"] = []
                    logger.info("Inicializando campo 'imagenes' en la fila")

                # Verificar si imagenes es un n√∫mero entero (contador) en lugar de una
                # lista
                if isinstance(fila.get("imagenes"), int):
                    logger.info(f"Campo 'imagenes' es un entero: {fila['imagenes']}")
                    # Buscar im√°genes reales en otros campos
                    if "imagen_data" in fila:
                        fila["imagenes"] = fila["imagen_data"]
                        logger.info(
                            f"Usando 'imagen_data' como fuente de im√°genes: {fila['imagenes']}"
                        )
                    else:
                        fila["imagenes"] = []
                        logger.info(
                            "No se encontraron im√°genes reales, inicializando lista vac√≠a"
                        )

                # Asegurarse de que imagenes sea una lista
                if not isinstance(fila["imagenes"], list):
                    fila["imagenes"] = [fila["imagenes"]] if fila["imagenes"] else []
                    logger.info(
                        f"Convertido campo 'imagenes' a lista: {fila['imagenes']}"
                    )

                if isinstance(imagenes_a_eliminar, list):
                    # Eliminar las im√°genes del servidor y de S3
                    if current_app.static_folder is None:
                        logger.error(
                            "Error de configuraci√≥n del servidor: static_folder es None"
                        )
                        flash("Error de configuraci√≥n del servidor", "error")
                        return redirect(url_for("main.tables"))
                    ruta_uploads = os.path.join(current_app.static_folder, "uploads")
                    use_s3 = os.environ.get("USE_S3", "false").lower() == "true"

                    for img_a_eliminar in imagenes_a_eliminar:
                        try:
                            # Eliminar de S3 si est√° habilitado
                            if use_s3:
                                from app.utils.s3_utils import delete_file_from_s3

                                delete_file_from_s3(img_a_eliminar)
                                logger.info(f"Imagen eliminada de S3: {img_a_eliminar}")

                            # Eliminar del servidor local
                            ruta_img = os.path.join(ruta_uploads, img_a_eliminar)
                            if os.path.exists(ruta_img):
                                os.remove(ruta_img)

                            # Limpiar cache de S3 para esta imagen
                            from app.utils.image_utils import clear_s3_cache

                            clear_s3_cache(img_a_eliminar)

                            logger.info(
                                f"Imagen eliminada del servidor local: {img_a_eliminar}"
                            )
                        except Exception as e:
                            logger.error(
                                f"Error al eliminar imagen: {str(e)}", exc_info=True
                            )

                    # Actualizar la lista de im√°genes en la fila
                    imagenes_actualizadas = [
                        img
                        for img in fila.get("imagenes", [])
                        if img not in imagenes_a_eliminar
                    ]
                    update_data[f"data.{fila_index}.imagenes"] = imagenes_actualizadas
                    logger.info(
                        f"Lista de im√°genes actualizada: {imagenes_actualizadas}"
                    )

                    # Actualizar tambi√©n el campo imagen_data si existe
                    if "imagen_data" in fila:
                        update_data[f"data.{fila_index}.imagen_data"] = (
                            imagenes_actualizadas
                        )
                        logger.info(
                            "Campo 'imagen_data' actualizado con la misma lista de im√°genes"
                        )

                    # Si no quedan im√°genes, establecer el contador a 0
                    if not imagenes_actualizadas and isinstance(
                        fila.get("imagenes"), int
                    ):
                        update_data[f"data.{fila_index}.imagenes"] = 0
                        logger.info("No quedan im√°genes, contador establecido a 0")
            except Exception as e:
                logger.error(f"Error al procesar im√°genes a eliminar: {str(e)}")

        # Procesar documentos a eliminar
        documentos_a_eliminar = request.form.get("documentos_a_eliminar", "")
        if documentos_a_eliminar and documentos_a_eliminar.strip():
            try:
                import json

                documentos_a_eliminar = json.loads(documentos_a_eliminar)
                # Solo procesar si realmente hay documentos para eliminar
                if documentos_a_eliminar and len(documentos_a_eliminar) > 0:
                    logger.info(f"Documentos a eliminar: {documentos_a_eliminar}")

                    # Eliminar los documentos de S3
                    use_s3 = os.environ.get("USE_S3", "false").lower() == "true"
                    if use_s3:
                        from app.utils.s3_utils import delete_file_from_s3

                        for doc_a_eliminar in documentos_a_eliminar:
                            try:
                                delete_file_from_s3(doc_a_eliminar)
                                logger.info(
                                    f"Documento eliminado de S3: {doc_a_eliminar}"
                                )
                            except Exception as e:
                                logger.error(
                                    f"Error eliminando documento de S3 {doc_a_eliminar}: {e}"
                                )

                    # Actualizar los documentos individuales en la fila
                    # Buscar campos Documentaci√≥n_0, Documentaci√≥n_1, Documentaci√≥n_2,
                    # etc.
                    for key, value in fila.items():
                        if key.startswith("Documentaci√≥n_") and isinstance(value, str):
                            if value in documentos_a_eliminar:
                                update_data[f"data.{fila_index}.{key}"] = ""
                                logger.info(
                                    f"Documento eliminado del campo {key}: {value}"
                                )

                    # Tambi√©n manejar el campo Documentaci√≥n original si existe
                    if "Documentaci√≥n" in fila and isinstance(
                        fila["Documentaci√≥n"], list
                    ):
                        documentos_actualizados = [
                            doc
                            for doc in fila["Documentaci√≥n"]
                            if doc not in documentos_a_eliminar
                        ]
                        update_data[f"data.{fila_index}.Documentaci√≥n"] = (
                            documentos_actualizados
                        )
                        logger.info(
                            f"Lista de documentos actualizada: {documentos_actualizados}"
                        )
                else:
                    logger.info(
                        "Campo documentos_a_eliminar est√° vac√≠o, no se eliminan documentos"
                    )

            except json.JSONDecodeError as e:
                logger.error(f"Error decodificando documentos_a_eliminar: {e}")
            except Exception as e:
                logger.error(f"Error procesando documentos a eliminar: {e}")
        else:
            logger.info("No se proporcion√≥ campo documentos_a_eliminar o est√° vac√≠o")

        # Procesar multimedia a eliminar
        multimedia_a_eliminar = request.form.get("multimedia_a_eliminar", "")
        if multimedia_a_eliminar:
            try:
                import json

                multimedia_a_eliminar = json.loads(multimedia_a_eliminar)
                logger.info(f"Multimedia a eliminar: {multimedia_a_eliminar}")

                if isinstance(multimedia_a_eliminar, list):
                    # Eliminar multimedia del servidor y de S3
                    use_s3 = os.environ.get("USE_S3", "false").lower() == "true"

                    for media_a_eliminar in multimedia_a_eliminar:
                        try:
                            # Eliminar de S3 si est√° habilitado
                            if use_s3:
                                from app.utils.s3_utils import delete_file_from_s3

                                delete_file_from_s3(media_a_eliminar)
                                logger.info(
                                    f"Multimedia eliminado de S3: {media_a_eliminar}"
                                )

                            # Limpiar cache de S3 para este multimedia
                            from app.utils.image_utils import clear_s3_cache

                            clear_s3_cache(media_a_eliminar)

                            # Eliminar del servidor local
                            if current_app.static_folder:
                                ruta_uploads = os.path.join(
                                    current_app.static_folder, "uploads"
                                )
                                ruta_media = os.path.join(
                                    ruta_uploads, media_a_eliminar
                                )
                                if os.path.exists(ruta_media):
                                    os.remove(ruta_media)
                                    logger.info(
                                        f"Multimedia eliminado del servidor local: {media_a_eliminar}"
                                    )
                        except Exception as e:
                            logger.error(
                                f"Error al eliminar multimedia: {str(e)}", exc_info=True
                            )

                    # Limpiar el campo multimedia en la fila
                    update_data[f"data.{fila_index}.Multimedia"] = ""
                    logger.info("Campo multimedia limpiado")

            except json.JSONDecodeError as e:
                logger.error(f"Error decodificando multimedia_a_eliminar: {e}")
            except Exception as e:
                logger.error(f"Error procesando multimedia a eliminar: {e}")

        # Procesar nuevas im√°genes si existen
        nuevas_imagenes = []
        if "images" in request.files:
            archivos = request.files.getlist("images")
            # Limitar a m√°ximo 3 im√°genes
            archivos = archivos[:3] if len(archivos) > 3 else archivos

            logger.info(f"Procesando {len(archivos)} nuevas im√°genes")

            for archivo in archivos:
                if archivo and archivo.filename and archivo.filename.strip():
                    # Generar nombre seguro y √∫nico para el archivo
                    nombre_seguro = secure_filename(archivo.filename)
                    extension = os.path.splitext(nombre_seguro)[1].lower()
                    nombre_unico = f"{uuid.uuid4().hex}{extension}"

                    # Verificar que sea una imagen v√°lida
                    if extension.lower() not in [".jpg", ".jpeg", ".png", ".gif"]:
                        logger.warning(f"Extensi√≥n no v√°lida para imagen: {extension}")
                        continue

                    # Guardar la imagen en la carpeta de uploads
                    if current_app.static_folder is None:
                        logger.error(
                            "Error de configuraci√≥n del servidor: static_folder es None"
                        )
                        continue
                    ruta_uploads = os.path.join(current_app.static_folder, "uploads")
                    if not os.path.exists(ruta_uploads):
                        os.makedirs(ruta_uploads)
                        logger.info(f"Carpeta de uploads creada: {ruta_uploads}")

                    ruta_completa = os.path.join(ruta_uploads, nombre_unico)
                    archivo.save(ruta_completa)
                    logger.info(f"Imagen guardada localmente: {ruta_completa}")

                    # Subir a S3 si est√° habilitado
                    use_s3 = os.environ.get("USE_S3", "false").lower() == "true"
                    logger.info(f"USE_S3: {use_s3}")
                    if use_s3:
                        try:
                            from app.utils.s3_utils import upload_file_to_s3

                            logger.info(f"Subiendo imagen a S3: {nombre_unico}")

                            result = upload_file_to_s3(ruta_completa, nombre_unico)
                            if result["success"]:
                                logger.info(f"Imagen subida a S3: {result['url']}")
                                # Eliminar el archivo local despu√©s de subirlo a S3
                                os.remove(ruta_completa)
                                logger.info(
                                    f"Imagen eliminada del servidor local despu√©s de subirla a S3: {nombre_unico}"
                                )
                            else:
                                logger.error(
                                    f"Error al subir imagen a S3: {result.get('error')}"
                                )
                        except Exception as e:
                            logger.error(
                                f"Error al procesar subida a S3: {str(e)}",
                                exc_info=True,
                            )

                    nuevas_imagenes.append(nombre_unico)

        # Si hay nuevas im√°genes, actualizar la lista en la base de datos
        if nuevas_imagenes:
            logger.info(f"Nuevas im√°genes a guardar: {nuevas_imagenes}")

            # Actualizar el campo images con las nuevas im√°genes
            if f"data.{fila_index}.images" in update_data:
                # Ya se actualiz√≥ por eliminaci√≥n, agregar las nuevas
                update_data[f"data.{fila_index}.images"] = (
                    update_data[f"data.{fila_index}.images"] + nuevas_imagenes
                )
                logger.info(
                    f"A√±adiendo nuevas im√°genes a la lista actualizada: {update_data[f'data.{fila_index}.images']}"
                )
            else:
                # Si ya hay im√°genes, a√±adir las nuevas
                if "images" in fila and isinstance(fila["images"], list):
                    update_data[f"data.{fila_index}.images"] = (
                        fila["images"] + nuevas_imagenes
                    )
                    logger.info(
                        f"A√±adiendo nuevas im√°genes a la lista existente: {fila['images'] + nuevas_imagenes}"
                    )
                else:
                    # Si no hay im√°genes previas, crear la lista
                    update_data[f"data.{fila_index}.images"] = nuevas_imagenes
                    logger.info(f"Creando nueva lista de im√°genes: {nuevas_imagenes}")

                # Actualizar tambi√©n imagen_data si existe
                if "imagen_data" in fila:
                    if isinstance(fila["imagen_data"], list):
                        update_data[f"data.{fila_index}.imagen_data"] = (
                            fila["imagen_data"] + nuevas_imagenes
                        )
                    else:
                        update_data[f"data.{fila_index}.imagen_data"] = nuevas_imagenes
                    logger.info(
                        "Actualizando campo imagen_data con las mismas im√°genes"
                    )

        # Preparar la actualizaci√≥n para MongoDB
        mongo_update = {}

        # Manejar los campos principales
        for key, value in update_data.items():
            if isinstance(value, dict):
                # Si es un diccionario, lo manejamos con $set en campos individuales
                for subkey, subvalue in value.items():
                    # Escapar los nombres de campos que tienen caracteres especiales
                    mongo_update[f"{key}.{subkey}"] = subvalue
            else:
                # Si no es un diccionario, lo manejamos directamente
                mongo_update[key] = value

        current_app.logger.info(f"Actualizando documento con datos: {mongo_update}")

        # Actualizar la fila en la base de datos
        g.spreadsheets_collection.update_one(
            {"_id": ObjectId(tabla_id)}, {"$set": mongo_update}
        )

        flash("Fila actualizada correctamente.", "success")

        # Calcular la p√°gina donde est√° la fila editada
        filas_por_pagina = 10
        pagina_fila = (fila_index // filas_por_pagina) + 1

        current_app.logger.info(f"[REDIRECT] Fila {fila_index} ‚Üí P√°gina {pagina_fila}")

        # Construir URL con par√°metros de forma m√°s robusta
        from urllib.parse import urlencode

        base_url = url_for("main.ver_tabla", table_id=tabla_id)
        params = {"page": pagina_fila}
        fragment = f"fila-{fila_index + 1}"  # Mostrar n√∫mero de fila desde 1
        redirect_url = f"{base_url}?{urlencode(params)}#{fragment}"

        current_app.logger.info(f"[REDIRECT] URL final: {redirect_url}")

        current_app.logger.info("[DEBUG_EDIT] Redirigiendo despu√©s de guardar")
        return redirect(redirect_url)

    current_app.logger.info("[DEBUG_EDIT] Renderizando template editar_fila.html")
    return render_template(
        "editar_fila.html",
        fila=fila,
        headers=headers,
        catalog=table_info,
        tabla_id=tabla_id,
        fila_index=fila_index,
    )


@main_bp.route("/agregar_fila/<tabla_id>", methods=["GET", "POST"])
def agregar_fila(tabla_id):
    from datetime import datetime

    current_app.logger.info(
        f"[AGREGAR_FILA] Iniciando agregar_fila para tabla_id: {tabla_id}, m√©todo: {request.method}"
    )
    print(
        f"[AGREGAR_FILA] PRINT: Iniciando agregar_fila para tabla_id: {tabla_id}, m√©todo: {request.method}"
    )
    print(f"[AGREGAR_FILA] PRINT: FUNCI√ìN AGREGAR_FILA EJECUT√ÅNDOSE")

    # Verificar sesi√≥n
    if not session.get("logged_in") or "username" not in session:
        flash("Debe iniciar sesi√≥n para agregar filas", "warning")
        return redirect(url_for("auth.login"))

    try:
        # Obtener la tabla
        tabla = g.spreadsheets_collection.find_one({"_id": ObjectId(tabla_id)})
        if not tabla:
            flash("Tabla no encontrada", "error")
            return redirect(url_for("main.dashboard_user"))

        # Verificar permisos: solo el propietario o admin puede agregar filas
        username = session.get("username")
        role = session.get("role", "user")

        # Obtener el propietario de la tabla (puede estar en diferentes campos)
        owner = tabla.get("owner") or tabla.get("created_by") or tabla.get("owner_name")

        current_app.logger.info(
            f"[DEBUG] Verificando permisos para agregar fila: role={role}, username={username}, owner={owner}"
        )

        if role != "admin" and owner != username:
            current_app.logger.warning(
                f"[DEBUG] Permiso denegado para agregar fila: {username} intentando modificar tabla de {owner}"
            )
            flash("No tiene permisos para agregar filas a esta tabla", "warning")
            return redirect(url_for("main.ver_tabla", table_id=tabla_id))

        if request.method == "POST":
            # Obtener datos del formulario
            nueva_fila = {}

            # PRE-PROCESAMIENTO: Manejar m√∫ltiples columnas de Documentaci√≥n
            # Buscar todos los campos Documentaci√≥n_file_INDEX y Documentaci√≥n_url_INDEX
            use_s3 = os.environ.get("USE_S3", "false").lower() == "true"

            for key, value in request.form.items():
                if key.startswith("Documentaci√≥n_file_") or key.startswith(
                    "Documentos_file_"
                ):
                    # Este campo ya se procesar√° en el bucle principal, saltar
                    continue
                elif key.startswith("Documentaci√≥n_url_") or key.startswith(
                    "Documentos_url_"
                ):
                    # Extraer el √≠ndice de la columna
                    if "Documentaci√≥n_url_" in key:
                        column_index = key.replace("Documentaci√≥n_url_", "")
                        header = "Documentaci√≥n"
                    else:
                        column_index = key.replace("Documentos_url_", "")
                        header = "Documentos"

                    if value and value.strip():
                        # Guardar URL directamente
                        field_name = f"{header}_{column_index}"
                        nueva_fila[field_name] = value.strip()
                        current_app.logger.info(
                            f"[AGREGAR_FILA] URL guardada: {field_name} = {value.strip()}"
                        )

            # Procesar archivos de documentos
            current_app.logger.info(
                f"[AGREGAR_FILA] Procesando {len(request.files)} archivos"
            )
            current_app.logger.info(
                f"[AGREGAR_FILA] Archivos recibidos: {list(request.files.keys())}"
            )
            print(f"[AGREGAR_FILA] PRINT: Procesando {len(request.files)} archivos")
            print(
                f"[AGREGAR_FILA] PRINT: Archivos recibidos: {list(request.files.keys())}"
            )

            for key, file in request.files.items():
                current_app.logger.info(
                    f"[AGREGAR_FILA] Procesando archivo: {key}, filename: {file.filename}"
                )
                # Detectar archivos de documentos con diferentes formatos
                is_document_file = (
                    key.startswith("Documentaci√≥n_file_")
                    or key.startswith("Documentos_file_")
                    or (key.startswith("Documentaci√≥n_") and key.endswith("_file"))
                )
                current_app.logger.info(
                    f"[AGREGAR_FILA] ¬øEs archivo de documento? {is_document_file} para {key}"
                )
                current_app.logger.info(
                    f"[AGREGAR_FILA] DEBUG: key.startswith('Documentaci√≥n_file_'): {key.startswith('Documentaci√≥n_file_')}"
                )
                current_app.logger.info(
                    f"[AGREGAR_FILA] DEBUG: key.startswith('Documentos_file_'): {key.startswith('Documentos_file_')}"
                )
                current_app.logger.info(
                    f"[AGREGAR_FILA] DEBUG: key.startswith('Documentaci√≥n_'): {key.startswith('Documentaci√≥n_')}"
                )
                current_app.logger.info(
                    f"[AGREGAR_FILA] DEBUG: key.endswith('_file'): {key.endswith('_file')}"
                )
                current_app.logger.info(
                    f"[AGREGAR_FILA] DEBUG: key.startswith('Documentaci√≥n_') and key.endswith('_file'): {key.startswith('Documentaci√≥n_') and key.endswith('_file')}"
                )

                if is_document_file:
                    if file and file.filename:
                        # Extraer el √≠ndice de la columna - manejar ambos formatos
                        current_app.logger.info(
                            f"[AGREGAR_FILA] Procesando documento: {key}"
                        )
                        if key.startswith("Documentaci√≥n_") and key.endswith("_file"):
                            # Formato: Documentaci√≥n_1_file, Documentaci√≥n_2_file, etc.
                            column_index = key.replace("Documentaci√≥n_", "").replace(
                                "_file", ""
                            )
                            header = "Documentaci√≥n"
                            current_app.logger.info(
                                f"[AGREGAR_FILA] Formato 1: {key} -> {header}_{column_index}"
                            )
                        elif "Documentaci√≥n_file_" in key:
                            # Formato: Documentaci√≥n_file_1, Documentaci√≥n_file_2, etc.
                            column_index = key.replace("Documentaci√≥n_file_", "")
                            header = "Documentaci√≥n"
                            current_app.logger.info(
                                f"[AGREGAR_FILA] Formato 2: {key} -> {header}_{column_index}"
                            )
                        else:
                            # Formato: Documentos_file_1, Documentos_file_2, etc.
                            column_index = key.replace("Documentos_file_", "")
                            header = "Documentos"
                            current_app.logger.info(
                                f"[AGREGAR_FILA] Formato 3: {key} -> {header}_{column_index}"
                            )

                        # Procesar archivo
                        filename = secure_filename(
                            f"{uuid.uuid4().hex}_{file.filename}"
                        )
                        upload_dir = get_upload_dir()
                        file_path = os.path.join(upload_dir, filename)
                        file.save(file_path)

                        # Subir a S3 si est√° habilitado
                        if use_s3:
                            try:
                                from app.utils.s3_utils import upload_file_to_s3

                                current_app.logger.info(
                                    f"[AGREGAR_FILA] Subiendo documento a S3: {filename}"
                                )
                                result = upload_file_to_s3(file_path, filename)
                                if result and result.get("success"):
                                    current_app.logger.info(
                                        f"[AGREGAR_FILA] Documento subido exitosamente a S3: {filename}"
                                    )
                                    # Eliminar archivo local despu√©s de subirlo a S3
                                    os.remove(file_path)
                                else:
                                    error_msg = (
                                        result.get("error", "Error desconocido")
                                        if result
                                        else "Resultado None"
                                    )
                                    current_app.logger.error(
                                        f"[AGREGAR_FILA] Error subiendo documento a S3: {error_msg}"
                                    )
                            except Exception as e:
                                current_app.logger.error(
                                    f"[AGREGAR_FILA] Error subiendo documento a S3 {filename}: {e}"
                                )

                        # Guardar en la fila con nombre √∫nico
                        field_name = f"{header}_{column_index}"
                        nueva_fila[field_name] = filename
                        current_app.logger.info(
                            f"[AGREGAR_FILA] Documento guardado: {field_name} = {filename}"
                        )
                        current_app.logger.info(
                            f"[AGREGAR_FILA] nueva_fila actual: {nueva_fila}"
                        )

            # PROCESAMIENTO PRINCIPAL: Manejar otros campos
            for header in tabla.get("headers", []):
                if header == "Fecha":
                    # Columna inteligente: asignar fecha actual si est√° vac√≠a
                    fecha_valor = request.form.get(header, "").strip()
                    if not fecha_valor:
                        fecha_valor = datetime.now().strftime("%Y-%m-%d")
                    nueva_fila[header] = fecha_valor
                elif header == "Multimedia":
                    # Manejar campo Multimedia
                    multimedia_url = request.form.get(f"{header}_url", "").strip()
                    multimedia_file = request.files.get(f"{header}_file")

                    if multimedia_url:
                        nueva_fila[header] = multimedia_url
                    elif multimedia_file and multimedia_file.filename:
                        # Procesar archivo multimedia
                        filename = secure_filename(
                            f"{uuid.uuid4().hex}_{multimedia_file.filename}"
                        )
                        upload_dir = get_upload_dir()
                        file_path = os.path.join(upload_dir, filename)
                        multimedia_file.save(file_path)
                        nueva_fila[header] = filename
                    else:
                        nueva_fila[header] = ""
                elif header in ["Documentos", "Documentaci√≥n"]:
                    # Saltar procesamiento de Documentaci√≥n base - ya se proces√≥ arriba
                    # Pero NO saltar Documentaci√≥n_1, Documentaci√≥n_2, etc.
                    if header == "Documentaci√≥n" or header == "Documentos":
                        continue
                else:
                    # NO sobrescribir campos de Documentaci√≥n que ya tienen valores
                    if (
                        header.startswith("Documentaci√≥n_")
                        and header in nueva_fila
                        and nueva_fila[header]
                    ):
                        # Ya tiene un valor de documento, no sobrescribir
                        current_app.logger.info(
                            f"[AGREGAR_FILA] Manteniendo valor existente para {header}: {nueva_fila[header]}"
                        )
                        continue
                    nueva_fila[header] = request.form.get(header, "")

            # Procesar im√°genes si existen
            imagenes = []
            if "images" in request.files:
                archivos = request.files.getlist("images")
                # Limitar a m√°ximo 3 im√°genes
                archivos = archivos[:3] if len(archivos) > 3 else archivos

                logger.info(f"Procesando {len(archivos)} im√°genes para nueva fila")

                for archivo in archivos:
                    if archivo and archivo.filename and archivo.filename.strip():
                        # Generar nombre seguro y √∫nico para el archivo
                        nombre_seguro = secure_filename(archivo.filename)
                        extension = os.path.splitext(nombre_seguro)[1].lower()
                        nombre_unico = f"{uuid.uuid4().hex}{extension}"

                        # Verificar que sea una imagen v√°lida
                        if extension.lower() not in [".jpg", ".jpeg", ".png", ".gif"]:
                            logger.warning(
                                f"Extensi√≥n no v√°lida para imagen: {extension}"
                            )
                            continue

                        # Guardar la imagen en la carpeta de uploads
                        if current_app.static_folder is None:
                            logger.error(
                                "Error de configuraci√≥n del servidor: static_folder es None"
                            )
                            continue
                        ruta_uploads = os.path.join(
                            current_app.static_folder, "uploads"
                        )
                        if not os.path.exists(ruta_uploads):
                            os.makedirs(ruta_uploads)
                            logger.info(f"Carpeta de uploads creada: {ruta_uploads}")

                        ruta_completa = os.path.join(ruta_uploads, nombre_unico)
                        archivo.save(ruta_completa)
                        logger.info(f"Imagen guardada localmente: {ruta_completa}")

                        # Subir a S3 si est√° habilitado
                        use_s3 = os.environ.get("USE_S3", "false").lower() == "true"
                        logger.info(f"USE_S3: {use_s3}")
                        if use_s3:
                            try:
                                from app.utils.s3_utils import upload_file_to_s3

                                logger.info(f"Subiendo imagen a S3: {nombre_unico}")
                                result = upload_file_to_s3(ruta_completa, nombre_unico)
                                if result["success"]:
                                    logger.info(f"Imagen subida a S3: {result['url']}")
                                    # Eliminar el archivo local despu√©s de subirlo a S3
                                    os.remove(ruta_completa)
                                    logger.info(
                                        f"Imagen eliminada del servidor local despu√©s de subirla a S3: {nombre_unico}"
                                    )
                                else:
                                    logger.error(
                                        f"Error al subir imagen a S3: {result.get('error')}"
                                    )
                            except Exception as e:
                                logger.error(
                                    f"Error al procesar subida a S3: {str(e)}",
                                    exc_info=True,
                                )

                        imagenes.append(nombre_unico)

                logger.info(f"Total de im√°genes procesadas: {len(imagenes)}")

            # Agregar las im√°genes a la fila
            if imagenes:
                nueva_fila["images"] = imagenes
                logger.info(f"Im√°genes agregadas a la nueva fila: {imagenes}")
                nueva_fila["num_imagenes"] = len(imagenes)

            # Agregar la fila a la tabla (actualizar tanto data como rows para
            # compatibilidad)
            current_app.logger.info(
                f"[AGREGAR_FILA] Guardando nueva fila en BD: {nueva_fila}"
            )
            result = g.spreadsheets_collection.update_one(
                {"_id": ObjectId(tabla_id)},
                {
                    "$push": {"data": nueva_fila, "rows": nueva_fila},
                    "$inc": {"num_rows": 1},
                    "$set": {"updated_at": datetime.utcnow()},
                },
            )

            current_app.logger.info(
                f"[AGREGAR_FILA] Resultado de la actualizaci√≥n: matched_count={result.matched_count}, modified_count={result.modified_count}"
            )

            if result.modified_count > 0:
                flash("Fila agregada correctamente", "success")
                current_app.logger.info(
                    "[AGREGAR_FILA] Fila agregada exitosamente, redirigiendo a ver_tabla"
                )
            else:
                flash("Error al agregar la fila", "error")
                current_app.logger.error("[AGREGAR_FILA] No se pudo agregar la fila")

            return redirect(url_for("main.ver_tabla", table_id=tabla_id))

        # Para peticiones GET, mostrar el formulario
        return render_template("agregar_fila.html", catalog=tabla)
    except Exception as e:
        logger.error(f"Error al agregar fila: {str(e)}", exc_info=True)
        flash(f"Error al agregar fila: {str(e)}", "error")
        return redirect(url_for("main.dashboard_user"))


@main_bp.route("/tables", methods=["GET", "POST"])
def tables():
    """Redirige las solicitudes de la antigua ruta /tables a la nueva ruta /catalogs.
    Esta funci√≥n mantiene la compatibilidad con enlaces antiguos.
    """

    flash("La funcionalidad de Tablas ha sido integrada en Cat√°logos", "info")

    # Redirigir a la lista de cat√°logos
    return redirect(url_for("catalogs.list"))

    # COMENTADO: Esta ruta estaba causando conflicto con la ruta principal /ver_tabla/<table_id>
    # @main_bp.route("/ver_tabla/<table_id>", methods=["GET"])
    # def ver_tabla_redirect(table_id):
    #     """Redirige las solicitudes de la antigua ruta /ver_tabla a la nueva ruta /catalogs/view.
    #     Esta funci√≥n mantiene la compatibilidad con enlaces antiguos.
    #     """
    #     current_app.logger.info(
    #         f"Redirigiendo desde /ver_tabla/{table_id} a /catalogs/view/{table_id}"
    #     )
    #     flash("La vista de Tablas ha sido integrada en Cat√°logos", "info")
    #
    #     # Redirigir a la vista de cat√°logo
    #     return redirect(url_for("catalogs.view", catalog_id=table_id))

    # COMENTADO: Conflicto con la funci√≥n editar_fila principal
    # @main_bp.route("/editar_fila/<tabla_id>/<int:fila_index>", methods=["GET", "POST"])
    # def editar_fila_redirect(tabla_id, fila_index):
    #     """Redirige las solicitudes de la antigua ruta /editar_fila a la nueva ruta /catalogs/edit_row.
    #     Esta funci√≥n mantiene la compatibilidad con enlaces antiguos.
    #     """
    #     current_app.logger.info(
    #         f"Redirigiendo desde /editar_fila/{tabla_id}/{fila_index} a /catalogs/edit_row/{tabla_id}/{fila_index}"
    #     )
    #     flash("La edici√≥n de filas ahora se realiza en Cat√°logos", "info")
    #
    #     # Redirigir a la edici√≥n de fila en cat√°logos
    #     return redirect(
    #         url_for("catalogs.edit_row", catalog_id=tabla_id, row_index=fila_index)
    #     )

    # M√©todo POST: Crear nueva tabla
    try:
        table_name = request.form.get("table_name", "").strip()
        import_file = request.files.get("import_table")

        # Validar nombre de tabla
        if not table_name:
            flash("El nombre de la tabla es obligatorio.", "error")
            return redirect(url_for("main.tables"))

        # Comprobar duplicados para el mismo usuario
        owner = session.get("username")
        if g.spreadsheets_collection.find_one({"owner": owner, "name": table_name}):
            flash("Ya existe una tabla con ese nombre.", "error")
            return redirect(url_for("main.tables"))

        # Procesar archivo importado o crear tabla nueva
        if import_file and import_file.filename:
            try:
                ext = os.path.splitext(import_file.filename)[1].lower()
                unique_id = uuid.uuid4().hex
                filename = f"table_{unique_id}{ext}"

                # Asegurarse de que existe la carpeta de uploads
                if "UPLOAD_FOLDER" not in current_app.config:
                    flash(
                        "Error en la configuraci√≥n del servidor. Contacte al administrador.",
                        "danger",
                    )
                    return redirect(url_for("main.tables"))

                filepath = os.path.join(current_app.config["UPLOAD_FOLDER"], filename)
                import_file.save(filepath)

                # Procesar archivo seg√∫n su tipo
                headers = None
                rows = []

                if ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
                    try:
                        wb = openpyxl.load_workbook(filepath)
                        hoja = wb.active
                        headers = list(
                            next(hoja.iter_rows(min_row=1, max_row=1, values_only=True))
                        )
                        for row in hoja.iter_rows(min_row=2, values_only=True):
                            if any(row):
                                rows.append(
                                    {
                                        h: (row[i] if i < len(row) else "")
                                        for i, h in enumerate(headers)
                                    }
                                )
                        wb.close()
                    except Exception as e:
                        flash(f"Error al leer el archivo Excel: {str(e)}", "error")
                        return redirect(url_for("main.tables"))
                elif ext == ".csv":
                    try:
                        with open(filepath, newline="", encoding="utf-8") as csvfile:
                            reader = csv.reader(csvfile)
                            headers = list(next(reader, None))
                            for row in reader:
                                if any(row):
                                    rows.append(
                                        {
                                            h: (row[i] if i < len(row) else "")
                                            for i, h in enumerate(headers)
                                        }
                                    )
                    except Exception as e:
                        flash(f"Error al leer el archivo CSV: {str(e)}", "error")
                        return redirect(url_for("main.tables"))
                else:
                    flash(
                        "Formato de archivo no soportado. Solo se permiten archivos .xlsx, .xlsm, .xltx, .xltm o .csv",
                        "error",
                    )
                    return redirect(url_for("main.tables"))

                if not headers or not any(headers):
                    flash(
                        "El archivo importado no contiene encabezados v√°lidos.", "error"
                    )
                    return redirect(url_for("main.tables"))

                # Insertar la tabla en la base de datos
                result = g.spreadsheets_collection.insert_one(
                    {
                        "owner": session.get("username"),
                        "name": table_name,
                        "filename": filename,
                        "headers": headers,
                        "created_at": datetime.utcnow(),
                        "created_by": session["username"],
                        "data": rows,
                        "num_rows": len(rows),
                    }
                )

                session["selected_headers"] = headers
                return redirect(
                    url_for("main.ver_tabla", table_id=str(result.inserted_id))
                )

            except Exception as e:
                logger.error(f"Error al procesar archivo: {str(e)}", exc_info=True)
                flash(f"Error al procesar el archivo: {str(e)}", "error")
                return redirect(url_for("main.tables"))
        else:
            # Crear tabla nueva desde encabezados ingresados manualmente
            headers_str = request.form.get("table_headers", "").strip()
            headers = (
                [h.strip() for h in headers_str.split(",") if h.strip()]
                if headers_str
                else ["N√∫mero", "Descripci√≥n", "Peso", "Valor"]
            )

            if not headers or not any(headers):
                flash("Debes ingresar al menos un encabezado.", "error")
                return redirect(url_for("main.tables"))

            try:
                file_id = secrets.token_hex(8)
                filename = f"table_{file_id}.xlsx"

                # Asegurarse de que existe la carpeta de uploads
                if "UPLOAD_FOLDER" not in current_app.config:
                    flash(
                        "Error en la configuraci√≥n del servidor. Contacte al administrador.",
                        "danger",
                    )
                    return redirect(url_for("main.tables"))

                filepath = os.path.join(current_app.config["UPLOAD_FOLDER"], filename)

                wb = Workbook()
                hoja = wb.active
                hoja.append(headers)
                wb.save(filepath)
                wb.close()

                result = g.spreadsheets_collection.insert_one(
                    {
                        "owner": session.get("username"),
                        "name": table_name,
                        "filename": filename,
                        "headers": headers,
                        "created_at": datetime.utcnow(),
                        "created_by": session["username"],
                        "data": [],
                        "num_rows": 0,
                    }
                )

                session["selected_headers"] = headers
                return redirect(
                    url_for("main.ver_tabla", table_id=str(result.inserted_id))
                )
            except Exception as e:
                logger.error(f"Error al crear tabla manual: {str(e)}", exc_info=True)
                flash(f"Error al crear la tabla: {str(e)}", "error")
                return redirect(url_for("main.tables"))
    except Exception as e:
        logger.error(f"Error general en tables: {str(e)}", exc_info=True)
        flash(
            "Error al procesar la solicitud. Por favor, int√©ntelo de nuevo.", "danger"
        )
        return redirect(url_for("main.tables"))


@main_bp.route("/editar_tabla/<id>", methods=["GET", "POST"])
def editar_tabla(id):
    # Verificar sesi√≥n
    if "username" not in session:
        flash("Debe iniciar sesi√≥n para realizar esta acci√≥n", "warning")
        return redirect(url_for("auth.login"))

    try:
        # Obtener la tabla
        table = g.spreadsheets_collection.find_one({"_id": ObjectId(id)})

        if not table:
            flash("Tabla no encontrada.", "error")
            return redirect(url_for("main.dashboard_user"))

        # Verificar permisos: solo el propietario puede editar la tabla
        if session.get("role") != "admin" and table.get("owner") != session.get(
            "username"
        ):
            flash("No tiene permisos para editar esta tabla.", "error")
            return redirect(url_for("main.dashboard_user"))

        if request.method == "POST":
            current_app.logger.info("üî•üî•üî• [EDITAR_TABLA] POST recibido")
            current_app.logger.info(f"[EDITAR_TABLA] Form data: {dict(request.form)}")

            # Obtener los datos del formulario
            new_name = request.form.get("name", "").strip()
            headers_str = request.form.get("headers", "").strip()
            nueva_miniatura = request.form.get("miniatura", "").strip()

            # Procesar archivo de miniatura si existe
            miniatura_file = request.files.get("miniatura_file")
            if miniatura_file and miniatura_file.filename:
                current_app.logger.info(
                    f"[EDITAR_TABLA] Procesando archivo de miniatura: {miniatura_file.filename}"
                )
                try:
                    # Generar nombre seguro y √∫nico para el archivo
                    filename = secure_filename(
                        f"{uuid.uuid4().hex}_{miniatura_file.filename}"
                    )
                    upload_dir = get_upload_dir()
                    file_path = os.path.join(upload_dir, filename)
                    miniatura_file.save(file_path)

                    # Subir a S3 si est√° habilitado
                    use_s3 = os.environ.get("USE_S3", "false").lower() == "true"
                    if use_s3:
                        try:
                            from app.utils.s3_utils import upload_file_to_s3

                            result = upload_file_to_s3(file_path, filename)
                            if result["success"]:
                                nueva_miniatura = result["url"]
                                # Eliminar el archivo local despu√©s de subirlo a S3
                                os.remove(file_path)
                                current_app.logger.info(
                                    f"[EDITAR_TABLA] Miniatura subida a S3: {nueva_miniatura}"
                                )
                            else:
                                current_app.logger.error(
                                    f"[EDITAR_TABLA] Error al subir miniatura a S3: {result.get('error')}"
                                )
                                # Usar URL local como fallback
                                nueva_miniatura = url_for(
                                    "static", filename=f"uploads/{filename}"
                                )
                        except Exception as e:
                            current_app.logger.error(
                                f"[EDITAR_TABLA] Error al procesar subida a S3: {str(e)}"
                            )
                            # Usar URL local como fallback
                            nueva_miniatura = url_for(
                                "static", filename=f"uploads/{filename}"
                            )
                    else:
                        # Usar URL local
                        nueva_miniatura = url_for(
                            "static", filename=f"uploads/{filename}"
                        )

                    current_app.logger.info(
                        f"[EDITAR_TABLA] Miniatura procesada: {nueva_miniatura}"
                    )
                except Exception as e:
                    current_app.logger.error(
                        f"[EDITAR_TABLA] Error al procesar archivo de miniatura: {str(e)}"
                    )
                    flash(
                        f"Error al procesar el archivo de miniatura: {str(e)}", "error"
                    )

            current_app.logger.info(
                f"[EDITAR_TABLA] Datos procesados: name='{new_name}', headers='{headers_str}', miniatura='{nueva_miniatura}'"
            )

            # Validar los datos
            if not new_name:
                flash("El nombre de la tabla no puede estar vac√≠o.", "error")
                return render_template("editar_tabla.html", table=table)

            # Procesar los encabezados
            new_headers = [h.strip() for h in headers_str.split(",") if h.strip()]

            if not new_headers:
                flash("Debe proporcionar al menos un encabezado.", "error")
                return render_template("editar_tabla.html", table=table)

            # Verificar si los encabezados han cambiado
            old_headers = table.get("headers", [])

            # Actualizar la tabla en la base de datos
            update_data = {"name": new_name, "headers": new_headers}

            # Procesar la miniatura personalizada
            if nueva_miniatura:
                update_data["miniatura"] = nueva_miniatura
                current_app.logger.info(
                    f"[MINIATURA_CUSTOM] Configurada miniatura personalizada: {nueva_miniatura}"
                )
            else:
                # Si se env√≠a vac√≠o, eliminar el campo de miniatura personalizada
                # (se usar√° la autom√°tica en dashboard_user)
                update_data["miniatura"] = ""
                current_app.logger.info(
                    "[MINIATURA_CUSTOM] Miniatura personalizada eliminada, se usar√° autom√°tica"
                )

            # üî• PRIMERO: Siempre actualizar nombre, headers y miniatura
            basic_update = {
                "name": new_name,
                "headers": new_headers,
                "miniatura": nueva_miniatura if nueva_miniatura else "",
            }
            g.spreadsheets_collection.update_one(
                {"_id": ObjectId(id)}, {"$set": basic_update}
            )
            current_app.logger.info(
                f"[UPDATE_BASIC] Actualizados: name={new_name}, headers={new_headers}, miniatura={nueva_miniatura}"
            )

            # Si los encabezados cambiaron, actualizar los datos
            if set(old_headers) != set(new_headers):
                # Crear un mapeo de viejos a nuevos encabezados para los que coinciden
                header_map = {}
                for old_h in old_headers:
                    if old_h in new_headers:
                        header_map[old_h] = old_h

                # Actualizar los datos con los nuevos encabezados
                data = table.get("data", [])
                new_data = []

                for row in data:
                    new_row = {}
                    # Mantener los campos que coinciden
                    for old_h, new_h in header_map.items():
                        if old_h in row:
                            new_row[new_h] = row[old_h]

                    # A√±adir campos para los nuevos encabezados
                    for h in new_headers:
                        if h not in new_row:
                            new_row[h] = ""

                    # Mantener campos especiales como 'imagenes'
                    if "imagenes" in row:
                        new_row["imagenes"] = row["imagenes"]

                    new_data.append(new_row)

                update_data["data"] = new_data

                # Manejar actualizaci√≥n de im√°genes
                nuevas_imagenes = request.files.getlist("imagenes")
                if nuevas_imagenes:
                    logger.info(f"Nuevas im√°genes a guardar: {nuevas_imagenes}")

                    # Verificar si el campo imagenes es un n√∫mero entero (contador)
                    if "imagenes" in row and isinstance(row["imagenes"], int):
                        # Si es un contador, actualizar el campo imagen_data
                        if "imagen_data" in row and isinstance(
                            row["imagen_data"], list
                        ):
                            # A√±adir las nuevas im√°genes a imagen_data
                            update_data["imagen_data"] = row["imagen_data"] + [
                                img.filename for img in nuevas_imagenes
                            ]
                            # Actualizar el contador
                            update_data["imagenes"] = len(row["imagen_data"]) + len(
                                nuevas_imagenes
                            )
                            logger.info(
                                f"Actualizando contador de im√°genes a {update_data['imagenes']}"
                            )
                        else:
                            # Crear el campo imagen_data
                            update_data["imagen_data"] = [
                                img.filename for img in nuevas_imagenes
                            ]
                            # Actualizar el contador
                            update_data["imagenes"] = len(nuevas_imagenes)
                            logger.info(
                                f"Creando campo imagen_data con {len(nuevas_imagenes)} im√°genes"
                            )
                    else:
                        # Si ya hay im√°genes, a√±adir las nuevas
                        if "imagenes" in row and isinstance(row["imagenes"], list):
                            update_data["imagenes"] = row["imagenes"] + [
                                img.filename for img in nuevas_imagenes
                            ]
                            logger.info(
                                f"A√±adiendo nuevas im√°genes a la lista existente: {update_data['imagenes']}"
                            )
                        else:
                            # Si no hay im√°genes previas, crear la lista
                            update_data["imagenes"] = [
                                img.filename for img in nuevas_imagenes
                            ]
                            logger.info(
                                f"Creando nueva lista de im√°genes: {nuevas_imagenes}"
                            )

                        # Actualizar tambi√©n imagen_data si existe
                        if "imagen_data" in row:
                            if isinstance(row["imagen_data"], list):
                                update_data["imagen_data"] = row["imagen_data"] + [
                                    img.filename for img in nuevas_imagenes
                                ]
                            else:
                                update_data["imagen_data"] = [
                                    img.filename for img in nuevas_imagenes
                                ]
                            logger.info(
                                "Actualizando campo imagen_data con las mismas im√°genes"
                            )

                # Actualizar la tabla
                g.spreadsheets_collection.update_one(
                    {"_id": ObjectId(id)}, {"$set": update_data}
                )

                # Los headers cambiaron, se actualiz√≥ todo
                pass

            # üéØ SIEMPRE redirigir despu√©s de actualizar (con o sin cambio de headers)
            flash("Tabla actualizada correctamente.", "success")
            current_app.logger.info("[EDITAR_TABLA] ‚úÖ Redirigiendo a dashboard_user")
            return redirect(url_for("main.dashboard_user"))

        # GET: Mostrar formulario de edici√≥n
        return render_template("editar_tabla.html", table=table)

    except Exception as e:
        logger.error(f"Error en editar_tabla: {str(e)}", exc_info=True)
        flash(f"Error al editar la tabla: {str(e)}", "error")
        return redirect(url_for("main.dashboard_user"))


@main_bp.route("/delete_row/<tabla_id>/<int:fila_index>", methods=["POST"])
def delete_row(tabla_id, fila_index):
    """Eliminar una fila espec√≠fica de una tabla"""
    current_app.logger.info(
        f"[DELETE_ROW] Eliminando fila {fila_index} de tabla {tabla_id}"
    )

    # Verificar sesi√≥n
    if "username" not in session:
        flash("Debe iniciar sesi√≥n para realizar esta acci√≥n", "warning")
        return redirect(url_for("auth.login"))

    # Obtener info de la tabla
    table_info = g.spreadsheets_collection.find_one({"_id": ObjectId(tabla_id)})
    if not table_info:
        flash("Tabla no encontrada.", "error")
        return redirect(url_for("main.tables"))

    # Verificar permisos: solo el propietario o admin puede eliminar filas
    username = session.get("username")
    role = session.get("role", "user")
    owner = (
        table_info.get("owner")
        or table_info.get("created_by")
        or table_info.get("owner_name")
    )

    if role != "admin" and owner != username:
        flash("No tienes permisos para eliminar filas de esta tabla.", "warning")
        return redirect(url_for("main.ver_tabla", table_id=tabla_id))

    # Sincronizar 'rows' y 'data'
    if "rows" in table_info and table_info["rows"] is not None:
        table_info["data"] = table_info["rows"]
    elif "data" in table_info and table_info["data"] is not None:
        table_info["rows"] = table_info["data"]
    else:
        table_info["data"] = []
        table_info["rows"] = []

    current_rows = table_info.get("data", [])
    current_app.logger.info(f"[DELETE_ROW] Filas actuales: {len(current_rows)}")

    # Verificar que el √≠ndice sea v√°lido
    if fila_index < 0 or fila_index >= len(current_rows):
        flash(f"√çndice de fila inv√°lido: {fila_index}.", "danger")
        current_app.logger.error(
            f"[DELETE_ROW] √çndice inv√°lido: {fila_index} >= {len(current_rows)}"
        )
        return redirect(url_for("main.ver_tabla", table_id=tabla_id))

    try:
        # Eliminar la fila
        deleted_row = current_rows.pop(fila_index)
        current_app.logger.info(f"[DELETE_ROW] Fila eliminada: {deleted_row}")

        # Actualizar en base de datos
        result = g.spreadsheets_collection.update_one(
            {"_id": ObjectId(tabla_id)},
            {
                "$set": {
                    "data": current_rows,
                    "rows": current_rows,
                    "num_rows": len(current_rows),
                    "updated_at": datetime.utcnow(),
                }
            },
        )

        if result.matched_count > 0 and result.modified_count > 0:
            flash("Fila eliminada correctamente", "success")
            current_app.logger.info(
                f"[DELETE_ROW] Fila eliminada exitosamente. Filas restantes: {len(current_rows)}"
            )
        else:
            flash("No se pudo eliminar la fila.", "warning")

    except Exception as e:
        current_app.logger.error(f"[DELETE_ROW] Error: {str(e)}")
        flash(f"Error al eliminar fila: {str(e)}", "danger")

    return redirect(url_for("main.ver_tabla", table_id=tabla_id))


@main_bp.route("/delete_table/<table_id>", methods=["POST"])
def delete_table(table_id):
    # Verificar sesi√≥n
    if "username" not in session:
        flash("Debe iniciar sesi√≥n para realizar esta acci√≥n", "warning")
        return redirect(url_for("auth.login"))

    # Verificar permisos: solo el propietario o admin puede eliminar la tabla
    username = session.get("username")
    role = session.get("role", "user")

    # Primero, obtener la tabla
    table = g.spreadsheets_collection.find_one({"_id": ObjectId(table_id)})

    if not table:
        flash("Tabla no encontrada.", "warning")
        return redirect(url_for("main.tables"))

    # Obtener el propietario de la tabla (puede estar en diferentes campos)
    owner = table.get("owner") or table.get("created_by") or table.get("owner_name")

    current_app.logger.info(
        f"[DEBUG] Verificando permisos para eliminar tabla: role={role}, username={username}, owner={owner}"
    )

    # Verificar permisos
    if role != "admin" and owner != username:
        current_app.logger.warning(
            f"[DEBUG] Permiso denegado para eliminar tabla: {username} intentando eliminar tabla de {owner}"
        )
        flash("No tiene permisos para eliminar esta tabla.", "warning")
        return redirect(url_for("main.tables"))

    if not table:
        flash("Tabla no encontrada o no tiene permisos para eliminarla.", "error")
        return redirect(url_for("main.tables"))

    # Eliminar archivo f√≠sico solo si existe el campo 'filename'
    filename = table.get("filename")
    if filename:
        filepath = os.path.join(current_app.config["UPLOAD_FOLDER"], filename)
        if os.path.exists(filepath):
            os.remove(filepath)

    g.spreadsheets_collection.delete_one({"_id": ObjectId(table_id)})

    if filename and session.get("selected_table") == filename:
        session.pop("selected_table", None)

    flash("Tabla eliminada exitosamente.", "success")
    return redirect(url_for("main.tables"))


@main_bp.route("/guia-rapida")
def guia_rapida():
    """Gu√≠a r√°pida de uso para usuarios normales"""
    return render_template("guia_rapida.html")


@main_bp.route("/soporte", methods=["GET", "POST"])
def soporte():
    if request.method == "POST":
        nombre = request.form.get("nombre", "")
        email = request.form.get("email", "")
        mensaje = request.form.get("mensaje", "")
        # Enviar email real a soporte
        subject = f"[Soporte] Mensaje de {nombre} ({email})"
        body_html = f"""
        <h3>Nuevo mensaje de soporte recibido</h3>
        <ul>
            <li><strong>Nombre:</strong> {nombre}</li>
            <li><strong>Email:</strong> {email}</li>
        </ul>
        <p><strong>Mensaje:</strong></p>
        <div style='white-space: pre-line; border:1px solid #eee; padding:10px; background:#f9f9f9;'>{mensaje}</div>
        <hr>
        <small>Este mensaje fue enviado desde el formulario de soporte de la aplicaci√≥n.</small>
        """
        # Destinatario: edfrutos@gmail.com
        recipients = ["edfrutos@gmail.com"]
        ok = notifications.send_email(subject, body_html, recipients)
        if ok:
            flash(
                "Tu mensaje ha sido enviado. El equipo de soporte te contactar√° pronto.",
                "success",
            )
        else:
            flash(
                "No se pudo enviar el mensaje. Intenta m√°s tarde o contacta por email.",
                "danger",
            )
        return redirect(url_for("main.guia_rapida"))
    return render_template("soporte.html")


@main_bp.route("/admin/catalogo/<catalog_id>/get-images", methods=["GET"])
@login_required
def get_catalog_images(catalog_id):
    """Obtener im√°genes del cat√°logo para selecci√≥n autom√°tica de miniatura"""
    try:
        # Obtener el cat√°logo
        catalog = g.spreadsheets_collection.find_one({"_id": ObjectId(catalog_id)})
        if not catalog:
            return {"error": "Cat√°logo no encontrado"}, 404

        # Verificar permisos: solo el propietario o admin puede acceder
        username = session.get("username")
        role = session.get("role", "user")
        owner = (
            catalog.get("owner")
            or catalog.get("created_by")
            or catalog.get("owner_name")
        )

        if role != "admin" and owner != username:
            return {"error": "No tienes permisos para acceder a este cat√°logo"}, 403

        # Buscar im√°genes en las filas del cat√°logo
        images = []
        data = catalog.get("data", [])

        for row in data:
            if not isinstance(row, dict):
                continue

            # Buscar en campos de im√°genes
            for field in ["imagenes", "images", "imagen_data", "Imagen"]:
                if field in row and row[field]:
                    if (
                        field == "Imagen"
                        and isinstance(row[field], str)
                        and row[field].startswith("http")
                    ):
                        # Imagen externa (Unsplash, etc.)
                        images.append(row[field])
                    elif isinstance(row[field], list):
                        # Lista de im√°genes
                        for img in row[field]:
                            if img and img != "N/A" and isinstance(img, str):
                                if img.startswith("http"):
                                    images.append(img)
                                else:
                                    # Imagen local
                                    use_s3 = (
                                        os.environ.get("USE_S3", "false").lower()
                                        == "true"
                                    )
                                    if use_s3:
                                        from app.utils.s3_utils import get_s3_url

                                        s3_url = get_s3_url(img)
                                        if s3_url:
                                            images.append(s3_url)
                                        else:
                                            images.append(
                                                url_for(
                                                    "static", filename=f"uploads/{img}"
                                                )
                                            )
                                    else:
                                        images.append(
                                            url_for("static", filename=f"uploads/{img}")
                                        )
                    elif isinstance(row[field], str) and row[field] != "N/A":
                        # Imagen individual
                        if row[field].startswith("http"):
                            images.append(row[field])
                        else:
                            # Imagen local
                            use_s3 = os.environ.get("USE_S3", "false").lower() == "true"
                            if use_s3:
                                from app.utils.s3_utils import get_s3_url

                                s3_url = get_s3_url(row[field])
                                if s3_url:
                                    images.append(s3_url)
                                else:
                                    images.append(
                                        url_for(
                                            "static", filename=f"uploads/{row[field]}"
                                        )
                                    )
                            else:
                                images.append(
                                    url_for("static", filename=f"uploads/{row[field]}")
                                )

        # Eliminar duplicados manteniendo el orden
        unique_images = []
        for img in images:
            if img not in unique_images:
                unique_images.append(img)

        return {"images": unique_images}

    except Exception as e:
        current_app.logger.error(f"Error al obtener im√°genes del cat√°logo: {str(e)}")
        return {"error": "Error interno del servidor"}, 500


@main_bp.route("/s3/<filename>")
def serve_s3_public(filename):
    """
    Ruta p√∫blica para servir archivos S3.
    Permite a usuarios normales acceder a archivos S3 sin autenticaci√≥n de admin.
    """
    current_app.logger.info(f"[S3-PUBLIC] üîç Solicitud para archivo: {filename}")

    try:
        # Configuraci√≥n S3 simplificada
        import boto3
        from botocore.exceptions import ClientError

        # Log de configuraci√≥n S3
        aws_key = current_app.config.get("AWS_ACCESS_KEY_ID")
        aws_secret = current_app.config.get("AWS_SECRET_ACCESS_KEY")
        aws_region = current_app.config.get("AWS_REGION", "eu-central-1")
        aws_bucket = current_app.config.get("S3_BUCKET_NAME")

        current_app.logger.info(
            f"[S3-PUBLIC] üîß Config S3 - Key: {'‚úÖ' if aws_key else '‚ùå'}, Secret: {'‚úÖ' if aws_secret else '‚ùå'}, Region: {aws_region}, Bucket: {aws_bucket}"
        )

        if not all([aws_key, aws_secret, aws_bucket]):
            current_app.logger.error("[S3-PUBLIC] ‚ùå Configuraci√≥n S3 incompleta")
            return "Configuraci√≥n S3 incompleta", 500

        # Crear cliente S3
        s3_client = boto3.client(
            "s3",
            aws_access_key_id=aws_key,
            aws_secret_access_key=aws_secret,
            region_name=aws_region,
        )

        # Intentar descargar el archivo desde S3
        try:
            response = s3_client.get_object(Bucket=aws_bucket, Key=filename)
            file_content = response["Body"].read()
            content_type = response.get("ContentType", "application/octet-stream")

            current_app.logger.info(
                f"[S3-PUBLIC] ‚úÖ Archivo descargado: {filename} ({len(file_content)} bytes)"
            )

            # Configurar Content-Type espec√≠fico para diferentes tipos de archivo
            if filename.lower().endswith(".pdf"):
                content_type = "application/pdf"
            elif filename.lower().endswith((".txt", ".md")):
                content_type = "text/plain; charset=utf-8"

            return Response(
                file_content,
                mimetype=content_type,
                headers={
                    "Content-Disposition": f"inline; filename={filename}",
                    "Cache-Control": "public, max-age=3600",
                },
            )

        except ClientError as e:
            error_code = e.response["Error"]["Code"]
            current_app.logger.error(
                f"[S3-PUBLIC] ‚ùå Error S3: {error_code} - {filename}"
            )

            if error_code == "NoSuchKey":
                return "Archivo no encontrado", 404
            else:
                return f"Error al acceder al archivo: {error_code}", 500

    except Exception as e:
        current_app.logger.error(
            f"[S3-PUBLIC] ‚ùå Error inesperado: {str(e)}", exc_info=True
        )
        return "Error interno del servidor", 500
