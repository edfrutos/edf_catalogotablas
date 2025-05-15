import os
from openpyxl import load_workbook, Workbook
from flask import session

def leer_datos_excel(filename):
    """Leer datos desde un archivo Excel"""
    if not os.path.exists(filename):
        return []

    wb = load_workbook(filename, read_only=True)
    hoja = wb.active
    data = []

    headers = [cell.value for cell in hoja[1]]

    for row in hoja.iter_rows(min_row=2, values_only=True):
        registro = {headers[i]: row[i] for i in range(len(headers))}
        # Asegurar que 'Imagenes' sea una lista
        if "Imagenes" in registro and isinstance(registro["Imagenes"], str):
            registro["Imagenes"] = registro["Imagenes"].split(", ")
        elif "Imagenes" in registro:
            registro["Imagenes"] = []

        data.append(registro)

    wb.close()
    return data

def escribir_datos_excel(data, filename):
    """Guardar datos en un archivo Excel"""
    wb = Workbook()
    hoja = wb.active
    hoja.title = "Datos"

    headers = session.get("selected_headers", ["Número", "Descripción", "Peso", "Valor", "Imagenes"])

    if "Número" not in headers:
        headers.insert(0, "Número")

    hoja.append(headers)

    for item in data:
        fila = []
        for header in headers:
            valor = item.get(header, "")
            if header == "Imagenes" and isinstance(valor, list):
                valor = ", ".join(valor)
            fila.append(valor)
        hoja.append(fila)

    wb.save(filename)
    wb.close()
