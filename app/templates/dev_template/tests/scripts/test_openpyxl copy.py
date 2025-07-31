import openpyxl
from openpyxl import Workbook

# Crear un nuevo libro de trabajo y una hoja
wb = Workbook()
ws = wb.active

# Escribir datos en la hoja
ws['A1'] = 'Hello'
ws['B1'] = 'World'

# Guardar el libro de trabajo
wb.save('test.xlsx')

print("Archivo Excel creado exitosamente.")